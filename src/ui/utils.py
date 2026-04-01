"""
ui/utils.py
===========
Standalone utility functions used across multiple screens.
"""

import os
import sys
from tkinter import filedialog, messagebox


def open_file(path: str):
    """Open a file with the system default application."""
    import subprocess
    path = str(path)
    if sys.platform == "darwin":
        subprocess.run(["open", path])
    elif sys.platform == "win32":
        os.startfile(path)
    else:
        subprocess.run(["xdg-open", path])


def export_to_excel(headers: list, rows: list, default_name: str):
    """Write headers + rows to an Excel file chosen by the user."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messagebox.showerror("Export", "openpyxl is not installed.\nRun: pip install openpyxl")
        return

    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile=default_name,
        title="Save Excel export",
    )
    if not path:
        return

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row styling
    header_fill = PatternFill("solid", fgColor="185FA5")  # BLUE
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, heading in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=heading)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data rows — alternate shading
    fill_alt = PatternFill("solid", fgColor="F0F0EE")
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if row_idx % 2 == 0:
                cell.fill = fill_alt

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"  # Keep header visible when scrolling

    wb.save(path)
    messagebox.showinfo("Export", f"Exported {len(rows)} rows to:\n{path}")

    # Open the file
    try:
        import subprocess, platform
        if platform.system() == "Darwin":
            subprocess.run(["open", path])
        elif platform.system() == "Windows":
            subprocess.run(["start", "", path], shell=True)
    except Exception:
        pass
