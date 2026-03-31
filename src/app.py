"""
app.py - Job Application Agent — Desktop GUI
=============================================
A clean, non-technical Mac desktop app built with Tkinter.
No command line needed. Double-click to launch.

Screens:
  1. Setup      — Upload CV + cover letter templates, configure basics
  2. Dashboard  — Overview of all applications with status badges
  3. Review     — Read each tailored application, approve or skip
  4. Settings   — Edit search keywords, API keys, salary threshold
"""

import tkinter as tk
import webbrowser
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import subprocess
import sys
import io
import re as _re
import json
import os
from pathlib import Path
from datetime import datetime


# ── Scan log capture ──────────────────────────────────────────────────────────

def _strip_ansi(text: str) -> str:
    """Remove ANSI / Rich escape sequences from a string."""
    return _re.sub(r"\x1b\[[0-9;]*[mGKHF]|\[/?[a-z_ ]+\]", "", text)


class _LogCapture(io.TextIOBase):
    """
    Drop-in stdout replacement that:
    - strips ANSI / Rich markup
    - appends plain lines to a list
    - calls an optional callback so the UI can update live
    - still writes to the original stdout so PyCharm/terminal stays intact
    """
    def __init__(self, real_stdout, on_line=None):
        self._real = real_stdout
        self._on_line = on_line      # callable(line: str) → None
        self._buf = ""
        self.lines: list[str] = []

    def write(self, text: str) -> int:
        self._real.write(text)       # pass through to terminal
        clean = _strip_ansi(text)
        self._buf += clean
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            if line.strip():
                self.lines.append(line)
                if self._on_line:
                    self._on_line(line)
        return len(text)

    def flush(self):
        self._real.flush()

# ── Colour palette ────────────────────────────────────────────────────────────
BG       = "#F7F6F3"
BG2      = "#EEECEA"
CARD     = "#FFFFFF"
BORDER   = "#D8D6D0"
TEXT     = "#1A1916"
TEXT2    = "#6B6963"
BLUE     = "#185FA5"
BLUE_LT  = "#E6F1FB"
GREEN    = "#3B6D11"
GREEN_LT = "#EAF3DE"
AMBER    = "#854F0B"
AMBER_LT = "#FAEEDA"
RED      = "#A32D2D"
RED_LT   = "#FCEBEB"
PURPLE   = "#534AB7"
PURP_LT  = "#EEEDFE"

STATUS_COLOURS = {
    "discovered":     (BLUE,   BLUE_LT),
    "score_me":       (AMBER,  AMBER_LT),
    "scored":         (GREEN,  GREEN_LT),
    "filtered":       (TEXT2,  BG2),
    "dismissed":      (TEXT2,  BG2),
    "tailoring":      (AMBER,  AMBER_LT),
    "tailored":       (BLUE,   BLUE_LT),
    "pending_review": (AMBER,  AMBER_LT),
    "approved":       (GREEN,  GREEN_LT),
    "skipped":        (RED,    RED_LT),
    "submitted":      (GREEN,  GREEN_LT),
    "interview":      (PURPLE, PURP_LT),
    "rejected":       (RED,    RED_LT),
    "offer":          (GREEN,  GREEN_LT),
}

FONT_FAMILY = "SF Pro Display" if sys.platform == "darwin" else "Segoe UI"
FONT        = (FONT_FAMILY, 13)
FONT_SM     = (FONT_FAMILY, 11)
FONT_LG     = (FONT_FAMILY, 16, "bold")
FONT_HEAD   = (FONT_FAMILY, 22, "bold")


def configure_styles():
    style = ttk.Style()
    style.theme_use("clam")
    style.configure("TFrame",       background=BG)
    style.configure("Card.TFrame",  background=CARD,  relief="flat")
    style.configure("TLabel",       background=BG,    foreground=TEXT, font=FONT)
    style.configure("Card.TLabel",  background=CARD,  foreground=TEXT, font=FONT)
    style.configure("Muted.TLabel", background=BG,    foreground=TEXT2, font=FONT_SM)
    style.configure("Head.TLabel",  background=BG,    foreground=TEXT, font=FONT_HEAD)
    style.configure("TButton",      font=FONT, padding=(12, 6))
    style.configure("Primary.TButton", background=BLUE, foreground="white", font=FONT)
    style.map("Primary.TButton", background=[("active", "#1050A0")])
    style.configure("TEntry",       font=FONT, padding=6)
    style.configure("TSeparator",   background=BORDER)
    style.configure("Treeview",     font=FONT_SM, rowheight=32, background=CARD,
                    fieldbackground=CARD, foreground=TEXT, borderwidth=0)
    style.configure("Treeview.Heading", font=(FONT_FAMILY, 11, "bold"),
                    background=BG2, foreground=TEXT2, relief="flat")
    style.map("Treeview", background=[("selected", BLUE_LT)], foreground=[("selected", BLUE)])


class JobAgentApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Job Application Agent")
        self.geometry("1100x720")
        self.minsize(900, 600)
        self.configure(bg=BG)
        configure_styles()

        # Load config + tracker lazily (may not be set up yet)
        self.config = None
        self.tracker = None
        self._load_backend()

        # Session activity log — records scoring/tailoring/apply events for email summary
        from session_log import SessionLog
        self._session = SessionLog()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        # ── First-run: prompt for output folder if path doesn't exist ────────
        self._check_output_dir_on_startup()

        # ── Layout: sidebar + content ────────────────────────────────────────
        self._sidebar = self._build_sidebar()
        self._sidebar.pack(side="left", fill="y")

        self._content = tk.Frame(self, bg=BG)
        self._content.pack(side="left", fill="both", expand=True)

        # ── Screens ──────────────────────────────────────────────────────────
        self._screens = {}
        self._current_screen = None
        self._scan_log_lines: list[str] = []   # persists across scans
        self._build_all_screens()

        # Show setup if not configured, else dashboard
        if self._needs_setup():
            self._show_screen("setup")
        else:
            self._show_screen("dashboard")

    def _on_close(self):
        """Send session summary email if there was any activity, then close."""
        if self._session.has_activity() and self.config:
            sent = self._session.send_summary(self.config)
            if sent:
                print("[Session Log] Summary email sent.")
        self.destroy()

    # ── Backend loading ───────────────────────────────────────────────────────

    def _load_backend(self):
        try:
            sys.path.insert(0, str(Path(__file__).parent))
            from config import Config
            from tracker import ApplicationTracker
            self.config = Config()
            self.tracker = ApplicationTracker(self.config.db_path)
            self._seed_bundled_files()
        except Exception as e:
            print(f"Backend load warning: {e}")

    def _seed_bundled_files(self):
        """When running as a packaged .app, copy bundled CV and template files to
        Application Support so they work on first launch without manual upload."""
        if not getattr(sys, "frozen", False):
            return  # Development mode — files live in project root

        import shutil
        meipass = Path(sys._MEIPASS)
        target_dir = Path.home() / "Library" / "Application Support" / "JobAgent"
        target_dir.mkdir(parents=True, exist_ok=True)

        for src_name, dest_name in [
            ("base_cv.md",                 "base_cv.md"),
            ("cv_template.docx",           "cv_template.docx"),
            ("cv_style.json",              "cv_style.json"),
            ("cover_letter_template.docx", "cover_letter_template.docx"),
            ("cover_letter_style.json",    "cover_letter_style.json"),
        ]:
            src  = meipass / src_name
            dest = target_dir / dest_name
            if src.exists() and not dest.exists():
                try:
                    shutil.copy2(src, dest)
                except Exception as e:
                    print(f"Could not seed {src_name}: {e}")

    def _needs_setup(self) -> bool:
        """True only if the app is completely unconfigured (no name set)."""
        if not self.config:
            return True
        return self.config.candidate_name in ("YOUR_NAME", "", None)

    def _check_output_dir_on_startup(self):
        """If the save folder doesn't exist, ask the user to choose one before continuing."""
        if not self.config:
            return
        output_dir = Path(self.config.output_dir)
        if output_dir.exists():
            return

        # Show a friendly first-run dialog
        dialog = tk.Toplevel(self)
        dialog.title("Choose a save folder")
        dialog.geometry("520x240")
        dialog.resizable(False, False)
        dialog.configure(bg=BG)
        dialog.grab_set()  # modal

        tk.Label(
            dialog,
            text="Welcome to Job Agent!",
            font=(FONT_FAMILY, 16, "bold"),
            bg=BG, fg=TEXT,
        ).pack(pady=(28, 4))

        tk.Label(
            dialog,
            text="Please choose a folder where your tailored CVs\nand cover letters will be saved.",
            font=FONT, bg=BG, fg=TEXT2, justify="center",
        ).pack(pady=(0, 16))

        path_var = tk.StringVar(value=str(Path.home() / "Documents" / "JobAgent"))
        path_row = tk.Frame(dialog, bg=BG)
        path_row.pack(fill="x", padx=32)
        path_entry = tk.Entry(path_row, textvariable=path_var, font=FONT, width=38)
        path_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))

        def browse():
            chosen = filedialog.askdirectory(title="Choose save folder", parent=dialog)
            if chosen:
                path_var.set(chosen)

        tk.Button(path_row, text="Browse…", font=FONT_SM, bg=BG2, fg=TEXT,
                  relief="flat", padx=8, command=browse).pack(side="left")

        def confirm():
            chosen = path_var.get().strip()
            if not chosen:
                messagebox.showwarning("Required", "Please choose a folder.", parent=dialog)
                return
            try:
                Path(chosen).mkdir(parents=True, exist_ok=True)
                self.config.output_dir = Path(chosen)
                self._save_output_dir_to_prefs(chosen)
                dialog.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"Could not create folder:\n{e}", parent=dialog)

        tk.Button(
            dialog, text="Save & continue",
            font=(FONT_FAMILY, 13, "bold"), bg=BLUE, fg="white",
            relief="flat", padx=16, pady=8,
            command=confirm,
        ).pack(pady=20)

        self.wait_window(dialog)

    def _save_output_dir_to_prefs(self, output_dir: str):
        """Persist OUTPUT_DIR to the user prefs file (survives app reinstalls)."""
        import re
        try:
            sys.path.insert(0, str(Path(__file__).parent))
            from config import USER_PREFS_PATH
            prefs_path = USER_PREFS_PATH
        except Exception:
            prefs_path = Path.home() / "Library" / "Application Support" / "JobAgent" / "user.env"

        prefs_path.parent.mkdir(parents=True, exist_ok=True)
        text = prefs_path.read_text(encoding="utf-8") if prefs_path.exists() else ""
        if "OUTPUT_DIR=" in text:
            text = re.sub(r"OUTPUT_DIR=.*", f"OUTPUT_DIR={output_dir}", text)
        else:
            text = text.rstrip("\n") + f"\nOUTPUT_DIR={output_dir}\n"
        prefs_path.write_text(text, encoding="utf-8")

    def _save_linkedin_creds_to_prefs(self, email: str, password: str):
        """Persist LinkedIn credentials to the user prefs file."""
        import re
        try:
            from config import USER_PREFS_PATH
            prefs_path = USER_PREFS_PATH
        except Exception:
            prefs_path = Path.home() / "Library" / "Application Support" / "JobAgent" / "user.env"

        prefs_path.parent.mkdir(parents=True, exist_ok=True)
        text = prefs_path.read_text(encoding="utf-8") if prefs_path.exists() else ""

        if email:
            if "LINKEDIN_APPLY_EMAIL=" in text:
                text = re.sub(r"LINKEDIN_APPLY_EMAIL=.*", f"LINKEDIN_APPLY_EMAIL={email}", text)
            else:
                text = text.rstrip("\n") + f"\nLINKEDIN_APPLY_EMAIL={email}\n"

        if password:
            if "LINKEDIN_APPLY_PASSWORD=" in text:
                text = re.sub(r"LINKEDIN_APPLY_PASSWORD=.*", f"LINKEDIN_APPLY_PASSWORD={password}", text)
            else:
                text = text.rstrip("\n") + f"\nLINKEDIN_APPLY_PASSWORD={password}\n"

        prefs_path.write_text(text, encoding="utf-8")

    def _save_email_prefs(self, notify_email: str, smtp_user: str, smtp_password: str):
        """Persist email notification settings to the user prefs file."""
        import re
        try:
            from config import USER_PREFS_PATH
            prefs_path = USER_PREFS_PATH
        except Exception:
            prefs_path = Path.home() / "Library" / "Application Support" / "JobAgent" / "user.env"

        prefs_path.parent.mkdir(parents=True, exist_ok=True)
        text = prefs_path.read_text(encoding="utf-8") if prefs_path.exists() else ""

        def set_pref(text, key, value):
            if not value:
                return text
            if f"{key}=" in text:
                return re.sub(rf"{key}=.*", f"{key}={value}", text)
            return text.rstrip("\n") + f"\n{key}={value}\n"

        text = set_pref(text, "NOTIFY_EMAIL",   notify_email)
        text = set_pref(text, "SMTP_USER",      smtp_user)
        text = set_pref(text, "SMTP_FROM",      smtp_user)
        text = set_pref(text, "SMTP_PASSWORD",  smtp_password)
        prefs_path.write_text(text, encoding="utf-8")

    # ── Sidebar ───────────────────────────────────────────────────────────────

    def _build_sidebar(self) -> tk.Frame:
        sidebar = tk.Frame(self, bg=BG, width=200)
        sidebar.pack_propagate(False)

        # Logo / title
        tk.Label(sidebar, text="Job Agent", font=(FONT_FAMILY, 17, "bold"),
                 bg=BG, fg=TEXT, pady=20).pack(fill="x", padx=20)

        tk.Frame(sidebar, bg=BORDER, height=1).pack(fill="x", padx=16)

        self._nav_buttons = {}
        nav_items = [
            ("dashboard", "  Dashboard"),
            ("screen",    "  Screen jobs"),
            ("review",    "  Review queue"),
            ("settings",  "  Settings"),
            ("setup",     "  Setup"),
            ("log",       "  Scan log"),
            ("database",  "  Database"),
            ("info",      "  Info"),
        ]
        for key, label in nav_items:
            btn = tk.Button(
                sidebar, text=label, anchor="w",
                font=FONT, bg=BG, fg=TEXT,
                activebackground=BG2, activeforeground=TEXT,
                relief="flat", bd=0, pady=10, padx=16,
                command=lambda k=key: self._show_screen(k)
            )
            btn.pack(fill="x")
            self._nav_buttons[key] = btn

        # Scan button at bottom
        tk.Frame(sidebar, bg=BG).pack(fill="both", expand=True)
        self._scan_btn = tk.Button(
            sidebar, text="  Scan for jobs",
            font=(FONT_FAMILY, 13, "bold"), anchor="w",
            bg=BLUE, fg=TEXT,
            activebackground="#1050A0", activeforeground=TEXT,
            relief="flat", bd=0, pady=12, padx=16,
            command=self._run_scan
        )
        self._scan_btn.pack(fill="x", padx=12, pady=12)

        return sidebar

    def _show_screen(self, name: str):
        if self._current_screen:
            self._screens[self._current_screen].pack_forget()
        # Highlight nav — selected = black bold, others = normal
        for key, btn in self._nav_buttons.items():
            if key == name:
                btn.config(bg=BG2, fg=TEXT,
                           font=(FONT_FAMILY, 13, "bold"))
            else:
                btn.config(bg=BG, fg=TEXT,
                           font=FONT)
        self._screens[name].pack(fill="both", expand=True)
        self._current_screen = name
        # Refresh data on switch
        if name == "dashboard":
            self._refresh_dashboard()
        elif name == "review":
            self._refresh_review()
        elif name == "screen":
            self._refresh_screen()
        elif name == "setup":
            self._refresh_settings()
        elif name == "log":
            self._refresh_log_screen()
        elif name == "database":
            self._db_refresh_all()
        elif name == "info":
            pass  # static screen, no refresh needed
        elif name == "settings":
            self._refresh_settings()

    # ── Screen builders ───────────────────────────────────────────────────────

    def _build_all_screens(self):
        for name, builder in [
            ("setup",     self._build_setup_screen),
            ("dashboard", self._build_dashboard_screen),
            ("screen",    self._build_screen_screen),
            ("review",    self._build_review_screen),
            ("settings",  self._build_settings_screen),
            ("log",       self._build_log_screen),
            ("database",  self._build_database_screen),
            ("info",      self._build_info_screen),
        ]:
            frame = tk.Frame(self._content, bg=BG)
            builder(frame)
            self._screens[name] = frame

    # ── SETUP SCREEN ──────────────────────────────────────────────────────────

    def _build_setup_screen(self, parent: tk.Frame):
        canvas = tk.Canvas(parent, bg=BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=BG)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        pad = dict(padx=40, pady=8)

        tk.Label(inner, text="Setup", font=FONT_HEAD, bg=BG, fg=TEXT).pack(anchor="w", padx=40, pady=(32, 4))
        tk.Label(inner, text="Connect your templates and API keys to get started.",
                 font=FONT_SM, bg=BG, fg=TEXT2).pack(anchor="w", **pad)

        # ── Section: Candidate ────────────────────────────────────────────────
        self._section_label(inner, "Candidate details")

        self._setup_name_var    = self._labeled_entry(inner, "Full name", "e.g. James Smith")
        self._setup_email_var   = self._labeled_entry(inner, "Email", "e.g. james@email.com")
        self._setup_phone_var   = self._labeled_entry(inner, "Phone", "+44 7XXX XXXXXX")
        self._setup_linkedin_var= self._labeled_entry(inner, "LinkedIn URL", "https://linkedin.com/in/...")

        self._setup_address_var  = self._labeled_entry(inner, "Address line 1", "e.g. 12 Example Street")
        self._setup_address2_var = self._labeled_entry(inner, "Address line 2 (city, postcode)", "e.g. London, SW1A 1AA")
        # Pre-fill from config
        if self.config:
            self._setup_name_var.set(self.config.candidate_name if self.config.candidate_name != "YOUR_SON_NAME" else "")
            self._setup_email_var.set(self.config.candidate_email if "@" in self.config.candidate_email else "")
            self._setup_phone_var.set(self.config.candidate_phone)
            self._setup_linkedin_var.set(self.config.candidate_linkedin if "linkedin" in self.config.candidate_linkedin else "")
            self._setup_address_var.set(getattr(self.config, "candidate_address", ""))
            self._setup_address2_var.set(getattr(self.config, "candidate_address2", ""))

        # ── Section: API Keys ─────────────────────────────────────────────────
        self._section_label(inner, "API keys")

        self._setup_anthropic_var = self._labeled_entry(inner, "Anthropic API key  (required for CV tailoring)",
                                                        "sk-ant-...", show="*")
        self._setup_reed_var      = self._labeled_entry(inner, "Reed API key  (optional — get free at reed.co.uk/developers)",
                                                        "xxxxxxxx-xxxx-...", show="*")

        if self.config:
            if self.config.anthropic_api_key not in ("YOUR_ANTHROPIC_API_KEY", ""):
                self._setup_anthropic_var.set(self.config.anthropic_api_key)
            if self.config.reed_api_key not in ("YOUR_REED_API_KEY", ""):
                self._setup_reed_var.set(self.config.reed_api_key)

        # ── Section: LinkedIn auto-apply ──────────────────────────────────────
        self._section_label(inner, "LinkedIn Easy Apply")
        tk.Label(inner, text="Used to log in when auto-applying to LinkedIn jobs.\nLeave blank if you don't want to use LinkedIn auto-apply.",
                 font=FONT_SM, bg=BG, fg=TEXT2, justify="left").pack(anchor="w", padx=40, pady=(4, 0))

        self._setup_li_email_var    = self._labeled_entry(inner, "LinkedIn email", "you@email.com")
        self._setup_li_password_var = self._labeled_entry(inner, "LinkedIn password", "", show="*")

        if self.config:
            if getattr(self.config, "linkedin_apply_email", ""):
                self._setup_li_email_var.set(self.config.linkedin_apply_email)
            if getattr(self.config, "linkedin_apply_password", ""):
                self._setup_li_password_var.set(self.config.linkedin_apply_password)

        # ── Section: Session summary email ───────────────────────────────────
        self._section_label(inner, "Session summary email (optional)")
        tk.Label(inner, text="Sends a summary of activity (costs, scored jobs, tailored CVs, applications) when you close the app.\nUse a Gmail App Password — myaccount.google.com/apppasswords",
                 font=FONT_SM, bg=BG, fg=TEXT2, justify="left").pack(anchor="w", padx=40, pady=(4, 0))

        self._setup_notify_email_var = self._labeled_entry(inner, "Send summary to (email)", "you@email.com")
        self._setup_smtp_user_var    = self._labeled_entry(inner, "Gmail address (sender)", "you@gmail.com")
        self._setup_smtp_pass_var    = self._labeled_entry(inner, "Gmail App Password", "", show="*")

        if self.config:
            if getattr(self.config, "notify_email", ""):
                self._setup_notify_email_var.set(self.config.notify_email)
            if getattr(self.config, "smtp_user", ""):
                self._setup_smtp_user_var.set(self.config.smtp_user)
            if getattr(self.config, "smtp_password", ""):
                self._setup_smtp_pass_var.set(self.config.smtp_password)

        # ── Section: Templates ────────────────────────────────────────────────
        self._section_label(inner, "Document templates (.docx)")
        tk.Label(inner, text="Upload the candidate's existing CV and cover letter.\n"
                             "The AI will match their style for every output document.",
                 font=FONT_SM, bg=BG, fg=TEXT2, justify="left").pack(anchor="w", **pad)

        self._cv_template_path = tk.StringVar()
        self._letter_template_path = tk.StringVar()

        if self.config:
            cv_style = self.config.base_dir / "cv_template.docx"
            letter_style = self.config.base_dir / "cover_letter_template.docx"
            if cv_style.exists():
                self._cv_template_path.set(str(cv_style))
            if letter_style.exists():
                self._letter_template_path.set(str(letter_style))

        self._file_picker_row(inner, "CV template (.docx)", self._cv_template_path, self._pick_cv_template)
        self._file_picker_row(inner, "Cover letter template (.docx)", self._letter_template_path, self._pick_letter_template)

        # ── Section: Output folder ────────────────────────────────────────────
        self._section_label(inner, "Output folder")
        tk.Label(inner, text="Where tailored CVs and cover letters will be saved.",
                 font=FONT_SM, bg=BG, fg=TEXT2, justify="left").pack(anchor="w", **pad)

        self._output_dir_var = tk.StringVar()
        if self.config:
            self._output_dir_var.set(str(self.config.output_dir))

        self._file_picker_row(inner, "Output folder", self._output_dir_var, self._pick_output_dir)

        # ── Save button ────────────────────────────────────────────────────────
        tk.Frame(inner, bg=BG, height=16).pack()
        save_btn = tk.Button(inner, text="Save and continue →",
                             font=(FONT_FAMILY, 13, "bold"),
                             bg=BLUE, fg=TEXT,
                             padx=24, pady=10,
                             command=self._save_setup)
        save_btn.pack(anchor="w", padx=40, pady=(0, 40))

    def _section_label(self, parent, text: str):
        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", padx=40, pady=(20, 0))
        tk.Label(parent, text=text, font=(FONT_FAMILY, 12, "bold"),
                 bg=BG, fg=TEXT2).pack(anchor="w", padx=40, pady=(8, 4))

    def _labeled_entry(self, parent, label: str, placeholder: str = "", show: str = "") -> tk.StringVar:
        tk.Label(parent, text=label, font=FONT_SM, bg=BG, fg=TEXT2).pack(anchor="w", padx=40, pady=(6, 0))
        var = tk.StringVar()
        entry_kwargs = dict(textvariable=var, font=FONT, width=52,
                            relief="solid", bd=1, bg=CARD)
        if show:
            entry_kwargs["show"] = show
        entry = tk.Entry(parent, **entry_kwargs)
        entry.pack(anchor="w", padx=40, pady=(2, 0), ipady=6)
        if placeholder and not var.get():
            entry.insert(0, placeholder)
            entry.config(fg=TEXT2)
            entry.bind("<FocusIn>", lambda e: (entry.delete(0, "end"), entry.config(fg=TEXT))
                       if entry.get() == placeholder else None)
        return var

    def _file_picker_row(self, parent, label: str, var: tk.StringVar, command):
        row = tk.Frame(parent, bg=BG)
        row.pack(anchor="w", padx=40, pady=4, fill="x")
        tk.Label(row, text=label, font=FONT_SM, bg=BG, fg=TEXT2, width=28, anchor="w").pack(side="left")
        tk.Entry(row, textvariable=var, font=FONT_SM, width=34,
                 relief="solid", bd=1, bg=CARD, state="readonly").pack(side="left", padx=(0, 8), ipady=5)
        tk.Button(row, text="Browse…", font=FONT_SM, command=command,
                  bg=BG2, fg=TEXT, relief="flat", padx=10, pady=4).pack(side="left")

    def _pick_cv_template(self):
        path = filedialog.askopenfilename(
            title="Select CV template",
            filetypes=[("Word documents", "*.docx"), ("All files", "*.*")]
        )
        if path:
            self._cv_template_path.set(path)

    def _pick_letter_template(self):
        path = filedialog.askopenfilename(
            title="Select cover letter template",
            filetypes=[("Word documents", "*.docx"), ("All files", "*.*")]
        )
        if path:
            self._letter_template_path.set(path)

    def _pick_output_dir(self):
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self._output_dir_var.set(path)

    def _save_setup(self):
        name    = self._setup_name_var.get().strip()
        email   = self._setup_email_var.get().strip()
        phone   = self._setup_phone_var.get().strip()
        linkedin= self._setup_linkedin_var.get().strip()
        address = self._setup_address_var.get().strip()
        address2= self._setup_address2_var.get().strip()
        anth    = self._setup_anthropic_var.get().strip()
        reed    = self._setup_reed_var.get().strip()
        li_email    = self._setup_li_email_var.get().strip()
        li_password = self._setup_li_password_var.get().strip()
        notify_email = self._setup_notify_email_var.get().strip()
        smtp_user    = self._setup_smtp_user_var.get().strip()
        smtp_pass    = self._setup_smtp_pass_var.get().strip()
        output_dir = self._output_dir_var.get().strip()

        if not name or name == "e.g. James Smith":
            messagebox.showwarning("Setup", "Please enter the candidate's full name.")
            return
        if not anth or anth.startswith("sk-ant-") is False and len(anth) < 10:
            if not messagebox.askyesno("Setup", "No Anthropic API key entered — CV tailoring will not work. Continue anyway?"):
                return

        # Apply output dir immediately to config
        if output_dir and self.config:
            from pathlib import Path as _Path
            self.config.output_dir = _Path(output_dir)
            self.config.output_dir.mkdir(parents=True, exist_ok=True)

        # Process uploaded templates
        self._process_templates()

        # Write config.py with updated values
        self._write_config(name=name, email=email, phone=phone,
                           linkedin=linkedin, address=address, address2=address2,
                           output_dir=output_dir, anthropic_key=anth, reed_key=reed)
        if li_email or li_password:
            self._save_linkedin_creds_to_prefs(li_email, li_password)
        if notify_email or smtp_user or smtp_pass:
            self._save_email_prefs(notify_email, smtp_user, smtp_pass)

        # Reload backend
        self._load_backend()

        messagebox.showinfo("Setup", "✓ Setup saved! You're ready to scan for jobs.")
        self._show_screen("dashboard")

    def _process_templates(self):
        """Extract style fingerprints from uploaded docx templates."""
        try:
            from document_processor import DocumentProcessor
            processor = DocumentProcessor()
            base_dir = Path(__file__).parent

            cv_path = self._cv_template_path.get()
            if cv_path and Path(cv_path).exists():
                # Copy template to project dir
                import shutil
                dest = base_dir / "cv_template.docx"
                shutil.copy2(cv_path, dest)
                # Extract text + style
                cv_text, cv_style = processor.extract_cv_template(dest)
                # Save text as base_cv.md
                (base_dir / "base_cv.md").write_text(cv_text, encoding="utf-8")
                # Save style fingerprint
                (base_dir / "cv_style.json").write_text(cv_style.to_json(), encoding="utf-8")

            letter_path = self._letter_template_path.get()
            if letter_path and Path(letter_path).exists():
                import shutil
                dest = base_dir / "cover_letter_template.docx"
                shutil.copy2(letter_path, dest)
                _, letter_style = processor.extract_cover_letter_template(dest)
                (base_dir / "cover_letter_style.json").write_text(letter_style.to_json(), encoding="utf-8")

        except Exception as e:
            messagebox.showwarning("Templates", f"Could not process templates: {e}\n\nYou can re-upload them later.")

    def _write_config(self, name, email, phone, linkedin, address="", address2="", output_dir="", anthropic_key="", reed_key=""):
        config_path = Path(__file__).parent / "config.py"
        if not config_path.exists():
            config_path = Path(__file__).parent.parent / "config.py"
        if not config_path.exists():
            return
        text = config_path.read_text(encoding="utf-8")

        def replace_val(text, attr, new_val):
            import re
            pattern = rf'({attr}: str = ")[^"]*(")'
            return re.sub(pattern, rf'\g<1>{new_val}\g<2>', text)

        def replace_path(text, attr, new_val):
            import re
            # Replace Path(...) default_factory patterns
            pattern = rf'({attr}.*?Path\.home\(\))[^)]*(\))'
            result = re.sub(pattern, rf'\g<1> / "{new_val}"\g<2>', text)
            if result == text:
                # Try simple string path pattern
                pattern2 = rf'({attr}.*?Path\()[^)]*(\))'
                result = re.sub(pattern2, rf'\g<1>"{new_val}"\g<2>', text)
            return result

        if name:
            text = replace_val(text, "candidate_name", name)
        if email:
            text = replace_val(text, "candidate_email", email)
        if phone:
            text = replace_val(text, "candidate_phone", phone)
        if linkedin:
            text = replace_val(text, "candidate_linkedin", linkedin)
        if address:
            text = replace_val(text, "candidate_address", address)
        if address2:
            text = replace_val(text, "candidate_address2", address2)
        if output_dir:
            # Write to user prefs file so it survives app reinstalls
            self._save_output_dir_to_prefs(output_dir)
        if anthropic_key:
            text = text.replace('os.getenv("ANTHROPIC_API_KEY", "YOUR_ANTHROPIC_API_KEY")', f'"{anthropic_key}"')
            text = replace_val(text, "anthropic_api_key", anthropic_key)
        if reed_key:
            text = text.replace('os.getenv("REED_API_KEY", "YOUR_REED_API_KEY")', f'"{reed_key}"')
            text = replace_val(text, "reed_api_key", reed_key)

        config_path.write_text(text, encoding="utf-8")

    # ── DASHBOARD SCREEN ──────────────────────────────────────────────────────

    def _build_dashboard_screen(self, parent: tk.Frame):
        # State
        self._dash_sort_col = "Found"
        self._dash_sort_asc = False
        self._dash_search_var = tk.StringVar()
        self._dash_all_jobs = []

        # Header
        header = tk.Frame(parent, bg=BG)
        header.pack(fill="x", padx=32, pady=(28, 12))
        tk.Label(header, text="Dashboard", font=FONT_HEAD, bg=BG, fg=TEXT).pack(side="left")
        tk.Button(header, text="Refresh", font=FONT_SM, bg=BG2, fg=TEXT,
                  relief="flat", padx=10, pady=4,
                  command=self._refresh_dashboard).pack(side="right")
        tk.Button(header, text="Export to Excel", font=FONT_SM, bg=GREEN_LT, fg=GREEN,
                  relief="flat", padx=10, pady=4,
                  command=self._export_dashboard).pack(side="right", padx=(0, 8))

        # Summary stat cards
        self._stat_frame = tk.Frame(parent, bg=BG)
        self._stat_frame.pack(fill="x", padx=32, pady=(0, 8))

        # Search bar
        search_bar = tk.Frame(parent, bg=BG)
        search_bar.pack(fill="x", padx=32, pady=(0, 6))
        tk.Label(search_bar, text="Search:", font=FONT_SM, bg=BG, fg=TEXT2).pack(side="left")
        dash_search = tk.Entry(search_bar, textvariable=self._dash_search_var,
                               font=FONT_SM, width=28, relief="solid", bd=1, bg=CARD)
        dash_search.pack(side="left", padx=(4, 4), ipady=3)
        self._dash_search_var.trace_add("write", lambda *_: self._apply_dash_filter())
        tk.Button(search_bar, text="✕", font=FONT_SM, bg=BG2, fg=TEXT2, relief="flat",
                  padx=4, command=lambda: self._dash_search_var.set("")).pack(side="left")

        # Applications table
        table_frame = tk.Frame(parent, bg=BG)
        table_frame.pack(fill="both", expand=True, padx=32)

        cols = ("ID", "Role", "Employer", "Match", "Status", "Found")
        self._tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        # Role and Employer stretch to fill available space; others are fixed.
        col_cfg = {
            "ID":       dict(width=40,  minwidth=30,  stretch=False),
            "Role":     dict(width=260, minwidth=120, stretch=True),
            "Employer": dict(width=160, minwidth=80,  stretch=True),
            "Match":    dict(width=60,  minwidth=50,  stretch=False),
            "Status":   dict(width=120, minwidth=80,  stretch=False),
            "Found":    dict(width=90,  minwidth=70,  stretch=False),
        }
        for col in cols:
            self._tree.heading(col, text=col,
                               command=lambda c=col: self._dash_sort(c))
            self._tree.column(col, **col_cfg[col])

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self._tree.bind("<Double-1>", self._on_tree_double_click)
        self._tree.bind("<<TreeviewSelect>>", self._on_tree_select)

        # Bottom action bar
        action_bar = tk.Frame(parent, bg=BG2)
        action_bar.pack(fill="x", side="bottom")
        tk.Label(action_bar, text="Double-click a row to open files · Right-click for status update",
                 font=FONT_SM, bg=BG2, fg=TEXT2).pack(side="left", padx=16, pady=8)

        self._build_context_menu()

    def _build_context_menu(self):
        self._ctx_menu = tk.Menu(self, tearoff=0, font=FONT_SM)
        for status in ["submitted", "interview", "offer", "rejected"]:
            self._ctx_menu.add_command(
                label=f"Mark as {status}",
                command=lambda s=status: self._update_selected_status(s)
            )
        self._ctx_menu.add_separator()
        self._ctx_menu.add_command(label="Open files", command=self._open_selected_files)
        self._tree.bind("<Button-2>", self._show_context_menu)   # Mac right-click
        self._tree.bind("<Control-Button-1>", self._show_context_menu)

    def _show_context_menu(self, event):
        row = self._tree.identify_row(event.y)
        if row:
            self._tree.selection_set(row)
            self._ctx_menu.post(event.x_root, event.y_root)

    def _refresh_dashboard(self):
        if not self.tracker:
            return

        # Rebuild stat cards
        for w in self._stat_frame.winfo_children():
            w.destroy()

        jobs = self.tracker.get_all_jobs()
        self._dash_all_jobs = jobs
        counts = {}
        for job in jobs:
            s = job["status"]
            counts[s] = counts.get(s, 0) + 1

        stat_order = ["tailored", "pending_review", "approved", "submitted", "interview", "offer"]
        labels = {"tailored": "To review", "pending_review": "Pending",
                  "approved": "Approved", "submitted": "Applied",
                  "interview": "Interviews", "offer": "Offers"}

        for status in stat_order:
            count = counts.get(status, 0)
            fg, bg = STATUS_COLOURS.get(status, (TEXT, BG2))
            card = tk.Frame(self._stat_frame, bg=bg, padx=16, pady=10)
            card.pack(side="left", padx=(0, 10), pady=4)
            tk.Label(card, text=str(count), font=(FONT_FAMILY, 22, "bold"),
                     bg=bg, fg=fg).pack()
            tk.Label(card, text=labels.get(status, status),
                     font=FONT_SM, bg=bg, fg=fg).pack()

        self._apply_dash_filter()

    def _dash_sort(self, col: str):
        if self._dash_sort_col == col:
            self._dash_sort_asc = not self._dash_sort_asc
        else:
            self._dash_sort_col = col
            self._dash_sort_asc = True
        # Update heading arrows
        col_map = {"ID": "id", "Role": "title", "Employer": "employer",
                   "Match": "match_score", "Status": "status", "Found": "date_found"}
        for c in ("ID", "Role", "Employer", "Match", "Status", "Found"):
            arrow = ""
            if c == col:
                arrow = " ▲" if self._dash_sort_asc else " ▼"
            self._tree.heading(c, text=c + arrow)
        self._apply_dash_filter()

    def _apply_dash_filter(self):
        if not hasattr(self, "_dash_all_jobs"):
            return
        query = self._dash_search_var.get().lower().strip()
        jobs = self._dash_all_jobs

        if query:
            jobs = [j for j in jobs if
                    query in (j.get("title") or "").lower() or
                    query in (j.get("employer") or "").lower() or
                    query in (j.get("status") or "").lower()]

        col_map = {"ID": "id", "Role": "title", "Employer": "employer",
                   "Match": "match_score", "Status": "status", "Found": "date_found"}
        if self._dash_sort_col and self._dash_sort_col in col_map:
            key = col_map[self._dash_sort_col]
            def sort_key(j):
                v = j.get(key)
                if v is None:
                    return (1, "")
                return (0, v) if isinstance(v, (int, float)) else (0, str(v).lower())
            jobs = sorted(jobs, key=sort_key, reverse=not self._dash_sort_asc)

        self._tree.delete(*self._tree.get_children())
        for job in jobs:
            score = f"{job['match_score']}%" if job.get("match_score") else "—"
            date = (job.get("date_found") or "")[:10]
            status = job.get("status", "")
            self._tree.insert("", "end",
                              iid=str(job["id"]),
                              values=(job["id"],
                                      job.get("title") or "—",
                                      job.get("employer") or "—",
                                      score,
                                      status.replace("_", " ").title(),
                                      date))

    def _export_dashboard(self):
        """Export the current dashboard view (filtered + sorted) to Excel."""
        if not hasattr(self, "_dash_all_jobs") or not self._dash_all_jobs:
            messagebox.showinfo("Export", "No data to export.")
            return

        # Collect currently visible rows from the treeview (respects filter/sort)
        rows = []
        for iid in self._tree.get_children():
            rows.append(self._tree.item(iid)["values"])

        headers = ["ID", "Role", "Employer", "Match", "Status", "Found"]
        self._export_to_excel(headers, rows, default_name="dashboard_export.xlsx")

    def _export_screen(self):
        """Export the current screen jobs view (filtered + sorted) to Excel."""
        if not hasattr(self, "_screen_all_jobs") or not self._screen_all_jobs:
            messagebox.showinfo("Export", "No data to export.")
            return

        # Apply same filters as the UI
        query = self._screen_search_var.get().lower().strip()
        high_only = getattr(self, "_screen_high_only", tk.BooleanVar()).get()
        jobs = self._screen_all_jobs

        if high_only:
            jobs = [j for j in jobs if j.get("match_score") is not None and j["match_score"] >= 70]
        if query:
            jobs = [j for j in jobs if
                    query in (j.get("title") or "").lower() or
                    query in (j.get("employer") or "").lower() or
                    query in (j.get("salary") or "").lower() or
                    query in (j.get("source") or "").lower()]

        headers = ["ID", "Role", "Employer", "Location", "Salary", "Match %", "Status", "Source", "Found", "URL"]
        rows = []
        for j in jobs:
            rows.append([
                j.get("id"),
                j.get("title") or "",
                j.get("employer") or "",
                j.get("location") or "",
                j.get("salary") or "",
                j.get("match_score") or "",
                (j.get("status") or "").replace("_", " ").title(),
                j.get("source") or "",
                (j.get("date_found") or "")[:10],
                j.get("url") or "",
            ])

        self._export_to_excel(headers, rows, default_name="screen_jobs_export.xlsx")

    def _export_to_excel(self, headers: list, rows: list, default_name: str):
        """Write headers + rows to an Excel file chosen by the user."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Export", "openpyxl is not installed.\nRun: pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_name,
            title="Save Excel export",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active

        # Header row styling
        header_fill = PatternFill("solid", fgColor="185FA5")  # BLUE
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col_idx, heading in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=heading)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Data rows — alternate shading
        fill_alt = PatternFill("solid", fgColor="F0F0EE")
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if row_idx % 2 == 0:
                    cell.fill = fill_alt

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"  # Keep header visible when scrolling

        wb.save(path)
        messagebox.showinfo("Export", f"Exported {len(rows)} rows to:\n{path}")

        # Open the file
        try:
            import subprocess, platform
            if platform.system() == "Darwin":
                subprocess.run(["open", path])
            elif platform.system() == "Windows":
                subprocess.run(["start", "", path], shell=True)
        except Exception:
            pass

    def _on_tree_select(self, event):
        pass  # Could show detail panel

    def _on_tree_double_click(self, event):
        self._open_selected_files()

    def _open_selected_files(self):
        selected = self._tree.selection()
        if not selected:
            messagebox.showinfo("No selection", "Select a row first.")
            return
        job_id = int(selected[0])
        job = self.tracker.get_job(job_id)
        if not job:
            return
        opened = False
        for path_key in ("tailored_cv_path", "cover_letter_path"):
            path = job.get(path_key)
            if path and Path(path).exists():
                self._open_file(path)
                opened = True
        if not opened:
            messagebox.showinfo("No files", "No tailored documents found for this job.\n\nScore and tailor it first.")

    def _update_selected_status(self, status: str):
        selected = self._tree.selection()
        if not selected:
            return
        job_id = int(selected[0])
        self.tracker.update_status(job_id, status, f"Updated via dashboard")
        self._refresh_dashboard()

    @staticmethod
    def _open_file(path: str):
        import subprocess, sys
        path = str(path)
        if sys.platform == "darwin":
            subprocess.run(["open", path])
        elif sys.platform == "win32":
            os.startfile(path)
        else:
            subprocess.run(["xdg-open", path])


    # ── SCREEN JOBS SCREEN ────────────────────────────────────────────────────

    def _build_screen_screen(self, parent: tk.Frame):
        """
        Screening panel — shows all scored jobs, lets user tick which to tailor.
        Only ticked jobs burn CV tailoring credits.
        """
        # State for sort/search/filter
        self._screen_all_jobs = []
        self._screen_sort_col = None
        self._screen_sort_asc = True
        self._screen_search_var = tk.StringVar()
        self._screen_high_only = tk.BooleanVar(value=False)

        # Header
        hdr = tk.Frame(parent, bg=BG)
        hdr.pack(fill="x", padx=32, pady=(28, 4))
        tk.Label(hdr, text="Screen jobs", font=FONT_HEAD, bg=BG, fg=TEXT).pack(side="left")
        self._screen_counter = tk.Label(hdr, text="", font=FONT_SM, bg=BG, fg=TEXT2)
        self._screen_counter.pack(side="left", padx=14)

        tk.Label(parent,
                 text="Tick the jobs you want to apply for, then click Score selected (~$0.01 each). No money spent until you click Score.",
                 font=FONT_SM, bg=BG, fg=TEXT2, justify="left").pack(anchor="w", padx=32, pady=(0,8))

        # ── Toolbar row 1: actions + search ──────────────────────────────────
        toolbar = tk.Frame(parent, bg=BG)
        toolbar.pack(fill="x", padx=32, pady=(0, 4))
        tk.Button(toolbar, text="Select all", font=(FONT_FAMILY, 11, "bold"), bg=BG2, fg=TEXT,
                  relief="flat", padx=10, pady=4, activebackground=BORDER, activeforeground=TEXT,
                  command=self._screen_select_all).pack(side="left", padx=(0, 6))
        tk.Button(toolbar, text="Deselect all", font=(FONT_FAMILY, 11, "bold"), bg=BG2, fg=TEXT,
                  relief="flat", padx=10, pady=4, activebackground=BORDER, activeforeground=TEXT,
                  command=self._screen_deselect_all).pack(side="left", padx=(0, 6))
        tk.Button(toolbar, text="Refresh", font=(FONT_FAMILY, 11, "bold"), bg=BG2, fg=TEXT,
                  relief="flat", padx=10, pady=4, activebackground=BORDER, activeforeground=TEXT,
                  command=self._refresh_screen).pack(side="left", padx=(0, 16))
        tk.Button(toolbar, text="Export to Excel", font=(FONT_FAMILY, 11, "bold"), bg=GREEN_LT, fg=GREEN,
                  relief="flat", padx=10, pady=4,
                  command=self._export_screen).pack(side="left", padx=(0, 6))

        tk.Label(toolbar, text="Search:", font=FONT_SM, bg=BG, fg=TEXT2).pack(side="left")
        search_entry = tk.Entry(toolbar, textvariable=self._screen_search_var,
                                font=FONT_SM, width=20, relief="solid", bd=1, bg=CARD)
        search_entry.pack(side="left", padx=(4, 4), ipady=3)
        self._screen_search_var.trace_add("write", lambda *_: self._apply_screen_filter())
        tk.Button(toolbar, text="✕", font=FONT_SM, bg=BG2, fg=TEXT2, relief="flat",
                  padx=4, command=lambda: self._screen_search_var.set("")).pack(side="left")

        # ── Toolbar row 2: filter chips ───────────────────────────────────────
        filter_bar = tk.Frame(parent, bg=BG)
        filter_bar.pack(fill="x", padx=32, pady=(0, 6))

        tk.Label(filter_bar, text="Filter:", font=FONT_SM, bg=BG, fg=TEXT2).pack(side="left", padx=(0, 6))

        self._high_only_btn = tk.Button(
            filter_bar, text="⭐  High ≥70% only",
            font=(FONT_FAMILY, 11, "bold"), bg=BG2, fg=TEXT,
            relief="flat", padx=10, pady=3,
            activebackground=GREEN_LT, activeforeground=GREEN,
            command=self._toggle_high_only
        )
        self._high_only_btn.pack(side="left", padx=(0, 6))

        self._show_filtered_var = tk.BooleanVar(value=False)
        tk.Checkbutton(filter_bar, text="Show filtered & hidden jobs",
                       variable=self._show_filtered_var,
                       font=FONT_SM, bg=BG, fg=TEXT2, activebackground=BG,
                       command=self._refresh_screen).pack(side="left", padx=(8, 0))

        # Scoring summary bar (populated by _update_screen_summary)
        self._screen_summary_bar = tk.Frame(parent, bg=BG2)
        self._screen_summary_bar.pack(fill="x", padx=32, pady=(0, 4))

        # Scrollable job list
        list_frame = tk.Frame(parent, bg=BG)
        list_frame.pack(fill="both", expand=True, padx=32)

        self._screen_canvas = tk.Canvas(list_frame, bg=BG, highlightthickness=0)
        vsb = ttk.Scrollbar(list_frame, orient="vertical", command=self._screen_canvas.yview)
        self._screen_inner = tk.Frame(self._screen_canvas, bg=BG)
        self._screen_inner.bind(
            "<Configure>",
            lambda e: self._screen_canvas.configure(scrollregion=self._screen_canvas.bbox("all"))
        )
        self._screen_canvas.create_window((0, 0), window=self._screen_inner, anchor="nw")
        self._screen_canvas.configure(yscrollcommand=vsb.set)
        self._screen_canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Mousewheel scrolling
        self._screen_canvas.bind_all("<MouseWheel>",
            lambda e: self._screen_canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        self._screen_vars = {}  # job_id → BooleanVar

        # Bottom action bar
        action_bar = tk.Frame(parent, bg=BG2)
        action_bar.pack(fill="x", side="bottom")
        self._score_btn = tk.Button(
            action_bar,
            text="📊  Score selected (~$0.01 each)",
            font=(FONT_FAMILY, 13, "bold"),
            bg=BLUE, fg=TEXT,
            padx=20, pady=10,
            command=self._score_selected
        )
        self._score_btn.pack(side="right", padx=(8, 16), pady=8)

        self._tailor_btn = tk.Button(
            action_bar,
            text="✍  Tailor scored jobs (~$0.05 each)",
            font=(FONT_FAMILY, 13, "bold"),
            bg=GREEN, fg="black", relief="flat",
            padx=20, pady=10,
            command=self._tailor_selected
        )
        self._tailor_btn.pack(side="right", padx=16, pady=8)
        self._screen_status_label = tk.Label(
            action_bar, text="", font=FONT_SM, bg=BG2, fg=TEXT2
        )
        self._screen_status_label.pack(side="left", padx=16, pady=8)


    def _bind_tooltip(self, widget, text: str):
        """Show a tooltip popup when hovering over a widget."""
        tip = None

        def show(event):
            nonlocal tip
            x = widget.winfo_rootx() + 20
            y = widget.winfo_rooty() + widget.winfo_height() + 4
            tip = tk.Toplevel(widget)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{x}+{y}")
            # Word-wrap at 400px
            lbl = tk.Label(tip, text=text, font=FONT_SM, bg="#FFFBE6",
                           fg=TEXT, relief="solid", bd=1,
                           wraplength=400, justify="left", padx=8, pady=6)
            lbl.pack()

        def hide(event):
            nonlocal tip
            if tip:
                tip.destroy()
                tip = None

        widget.bind("<Enter>", show)
        widget.bind("<Leave>", hide)

    def _update_screen_summary(self, all_db_jobs: list = None):
        """Rebuild the scoring summary bar with score distribution stats."""
        if not hasattr(self, "_screen_summary_bar"):
            return
        for w in self._screen_summary_bar.winfo_children():
            w.destroy()

        if all_db_jobs is None and self.tracker:
            all_db_jobs = self.tracker.get_all_jobs()
        if not all_db_jobs:
            return

        # Only include jobs currently in 'scored' status for high/medium counts.
        # Jobs that have moved on (tailoring, approved, etc.) keep their match_score
        # but must not be double-counted here.
        scored_active = [j for j in all_db_jobs
                         if j["status"] == "scored" and j.get("match_score") is not None]
        # Low-score jobs are auto-moved to 'filtered' during scoring (they have match_score set)
        scored_low    = [j for j in all_db_jobs
                         if j["status"] == "filtered" and j.get("match_score") is not None]
        dismissed     = sum(1 for j in all_db_jobs if j["status"] == "dismissed")

        if not scored_active and not scored_low:
            return  # No scoring done yet — hide the bar

        min_score = getattr(self.config, "min_match_score", 65) if self.config else 65
        high   = sum(1 for j in scored_active if j["match_score"] >= 70)
        medium = sum(1 for j in scored_active if min_score <= j["match_score"] < 70)
        low    = len(scored_low)

        bar = self._screen_summary_bar
        tk.Label(bar, text="  Scoring summary:",
                 font=(FONT_FAMILY, 10, "bold"), bg=BG2, fg=TEXT2).pack(side="left", padx=(4, 8))

        if high:
            tk.Label(bar, text=f"● {high} high (≥70%)",
                     font=(FONT_FAMILY, 10), bg=BG2, fg=GREEN).pack(side="left", padx=(0, 10))
        if medium:
            tk.Label(bar, text=f"● {medium} medium ({min_score}–69%)",
                     font=(FONT_FAMILY, 10), bg=BG2, fg=AMBER).pack(side="left", padx=(0, 10))
        if low:
            tk.Label(bar, text=f"● {low} low (<{min_score}%) — auto-hidden",
                     font=(FONT_FAMILY, 10), bg=BG2, fg=RED).pack(side="left", padx=(0, 10))
        if dismissed:
            tk.Label(bar, text=f"● {dismissed} dismissed by you",
                     font=(FONT_FAMILY, 10), bg=BG2, fg=TEXT2).pack(side="left", padx=(0, 10))
        if low or dismissed:
            tk.Label(bar, text="— tick 'Show filtered & hidden' to reveal",
                     font=(FONT_FAMILY, 10), bg=BG2, fg=TEXT2).pack(side="left")

    def _dismiss_job(self, job_id: int):
        """Mark a job as dismissed — hides it from the unfiltered view permanently."""
        if not self.tracker:
            return
        self.tracker.update_status(job_id, "dismissed", "Dismissed by user in Screen Jobs")
        self._screen_all_jobs = [j for j in self._screen_all_jobs if j["id"] != job_id]
        if job_id in self._screen_vars:
            del self._screen_vars[job_id]
        self._apply_screen_filter()
        self._update_screen_summary()

    def _refresh_screen(self):
        if not self.tracker:
            return

        # Clear existing rows and show loading splash
        for w in self._screen_inner.winfo_children():
            w.destroy()
        self._screen_vars.clear()

        loading = tk.Label(self._screen_inner, text="⏳  Loading jobs…",
                           font=FONT_LG, bg=BG, fg=TEXT2)
        loading.pack(pady=60)
        self._screen_inner.update_idletasks()
        self._screen_canvas.update_idletasks()

        show_filtered = self._show_filtered_var.get()

        all_db_jobs = self.tracker.get_all_jobs()
        if show_filtered:
            jobs = [j for j in all_db_jobs if j["status"] in
                    ("discovered", "score_me", "scored", "filtered", "dismissed")]
        else:
            jobs = [j for j in all_db_jobs if j["status"] in
                    ("discovered", "score_me", "scored")]

        discovered = sum(1 for j in jobs if j["status"] == "discovered")
        scored = sum(1 for j in jobs if j["status"] == "scored")
        self._screen_counter.config(text=f"{len(jobs)} jobs — {discovered} to review, {scored} scored")
        self._update_screen_summary(all_db_jobs)

        # Clear loading splash now data is ready
        for w in self._screen_inner.winfo_children():
            w.destroy()

        if not jobs:
            tk.Label(self._screen_inner,
                     text="No jobs yet. Click 'Scan for jobs' to start.",
                     font=FONT_SM, bg=BG, fg=TEXT2).pack(pady=40)
            return

        # Column headers — clickable to sort
        col_hdr = tk.Frame(self._screen_inner, bg=BG2)
        col_hdr.pack(fill="x", pady=(0, 2))
        tk.Label(col_hdr, text="", width=3, bg=BG2).pack(side="left")

        col_defs = [
            ("Role",     "title",        30),
            ("Employer", "employer",     18),
            ("Match",    "match_score",   7),
            ("Salary",   "salary",       14),
            ("Source",   "source",       10),
        ]
        for label, col_key, width in col_defs:
            def make_sort(k=col_key):
                def _sort():
                    if self._screen_sort_col == k:
                        self._screen_sort_asc = not self._screen_sort_asc
                    else:
                        self._screen_sort_col = k
                        self._screen_sort_asc = True
                    self._apply_screen_filter()
                return _sort
            arrow = ""
            if self._screen_sort_col == col_key:
                arrow = " ▲" if self._screen_sort_asc else " ▼"
            btn = tk.Button(col_hdr, text=label + arrow, width=width, anchor="w",
                            font=(FONT_FAMILY, 11, "bold"), bg=BG2, fg=TEXT2,
                            relief="flat", bd=0, activebackground=BORDER,
                            command=make_sort())
            btn.pack(side="left")

        tk.Frame(self._screen_inner, bg=BORDER, height=1).pack(fill="x")

        self._screen_all_jobs = jobs
        self._apply_screen_filter()

    def _apply_screen_filter(self):
        """Filter and sort _screen_all_jobs, then redraw rows."""
        if not hasattr(self, "_screen_all_jobs"):
            return
        query = self._screen_search_var.get().lower().strip()
        jobs = self._screen_all_jobs

        # High-score filter: only show jobs scored ≥70%
        if getattr(self, "_screen_high_only", tk.BooleanVar()).get():
            jobs = [j for j in jobs if j.get("match_score") is not None and j["match_score"] >= 70]

        if query:
            jobs = [j for j in jobs if
                    query in (j.get("title") or "").lower() or
                    query in (j.get("employer") or "").lower() or
                    query in (j.get("salary") or "").lower() or
                    query in (j.get("source") or "").lower()]

        if self._screen_sort_col:
            def sort_key(j):
                v = j.get(self._screen_sort_col)
                if v is None:
                    return (1, "")
                return (0, v) if isinstance(v, (int, float)) else (0, str(v).lower())
            jobs = sorted(jobs, key=sort_key, reverse=not self._screen_sort_asc)

        existing_checks = {jid: var.get() for jid, var in self._screen_vars.items()}

        children = self._screen_inner.winfo_children()
        for w in children[2:]:
            w.destroy()
        self._screen_vars.clear()

        if not jobs:
            high_only = getattr(self, "_screen_high_only", tk.BooleanVar()).get()
            msg = ("No high-scoring jobs (≥70%) found yet. Score some jobs first, or turn off the filter."
                   if high_only else "No jobs match your search.")
            tk.Label(self._screen_inner, text=msg,
                     font=FONT_SM, bg=BG, fg=TEXT2).pack(pady=20)
        else:
            for job in jobs:
                self._add_screen_row(job)
                if job["id"] in existing_checks:
                    self._screen_vars[job["id"]].set(existing_checks[job["id"]])

        # Update sort arrows
        col_defs = [("Role","title",30),("Employer","employer",18),
                    ("Match","match_score",7),("Salary","salary",14),("Source","source",10)]
        if children:
            col_hdr = children[0]
            hdr_btns = [w for w in col_hdr.winfo_children() if isinstance(w, tk.Button)]
            for btn, (label, col_key, _) in zip(hdr_btns, col_defs):
                arrow = ""
                if self._screen_sort_col == col_key:
                    arrow = " ▲" if self._screen_sort_asc else " ▼"
                btn.config(text=label + arrow)

        self._screen_status_label.config(
            text=f"{sum(v.get() for v in self._screen_vars.values())} selected"
        )

    def _add_screen_row(self, job: dict):
        job_id  = job["id"]
        score   = job.get("match_score")
        status  = job.get("status", "")
        is_filtered = status in ("filtered", "dismissed")

        row_bg = BG if not is_filtered else BG2

        row = tk.Frame(self._screen_inner, bg=row_bg, pady=3)
        row.pack(fill="x")

        # ── Pack side="right" widgets FIRST so they reserve space before left-packed widgets ──
        if not is_filtered:
            hide_btn = tk.Button(
                row, text="✕ Hide",
                font=(FONT_FAMILY, 9, "bold"), bg=RED_LT, fg=RED,
                relief="flat", padx=8, pady=2,
                activebackground="#F5CCCC", activeforeground=RED, cursor="hand2",
                command=lambda jid=job_id: self._dismiss_job(jid)
            )
            hide_btn.pack(side="right", padx=(0, 8))

        # ── Now pack left-side widgets in order ──────────────────────────────

        # Checkbox
        var = tk.BooleanVar(value=False)
        self._screen_vars[job_id] = var
        cb = tk.Checkbutton(row, variable=var, bg=row_bg,
                            activebackground=row_bg,
                            command=self._update_screen_count)
        cb.pack(side="left", padx=(4, 0))

        # Score badge colour
        if score is None:
            score_text = "—"
            score_fg, score_bg = TEXT2, BG2
        elif score >= 70:
            score_text = f"{score}%"
            score_fg, score_bg = GREEN, GREEN_LT
        elif score >= 55:
            score_text = f"{score}%"
            score_fg, score_bg = AMBER, AMBER_LT
        else:
            score_text = f"{score}%"
            score_fg, score_bg = RED, RED_LT

        title    = (job.get("title") or "—")[:48]
        employer = (job.get("employer") or "—")[:28]
        salary   = (job.get("salary") or "—")[:22]
        source   = (job.get("source") or "—").replace("_", " ")[:16]
        url      = job.get("url") or ""

        title_fg   = TEXT2 if is_filtered else BLUE
        title_font = (FONT_FAMILY, 11, "underline") if url and not is_filtered else FONT_SM

        title_lbl = tk.Label(row, text=title, width=30, anchor="w",
                             font=title_font, bg=row_bg, fg=title_fg, cursor="hand2" if url else "")
        title_lbl.pack(side="left")
        if url and not is_filtered:
            title_lbl.bind("<Button-1>", lambda e, u=url: webbrowser.open(u))
            title_lbl.bind("<Enter>", lambda e, l=title_lbl: l.config(fg=RED))
            title_lbl.bind("<Leave>", lambda e, l=title_lbl: l.config(fg=title_fg))

        tk.Label(row, text=employer, width=18, anchor="w", font=FONT_SM, bg=row_bg, fg=TEXT2).pack(side="left")

        # Score badge — shows reason on hover
        reason_text = (job.get("match_reason") or "").strip()
        badge = tk.Label(row, text=score_text, font=(FONT_FAMILY, 10, "bold"),
                         bg=score_bg, fg=score_fg, padx=6, pady=1,
                         cursor="question_arrow" if reason_text else "")
        badge.pack(side="left", padx=(0, 8))
        if reason_text:
            self._bind_tooltip(badge, reason_text)

        tk.Label(row, text=salary, width=14, anchor="w", font=FONT_SM, bg=row_bg, fg=TEXT2).pack(side="left")
        tk.Label(row, text=source, width=10, anchor="w", font=FONT_SM, bg=row_bg, fg=TEXT2).pack(side="left")

        if is_filtered:
            reason = (job.get("match_reason") or status)[:40]
            tk.Label(row, text=f"[{reason}]", font=(FONT_FAMILY, 10), bg=row_bg, fg=TEXT2).pack(side="left", padx=8)

        # Separator
        tk.Frame(self._screen_inner, bg=BORDER, height=1).pack(fill="x")

    def _toggle_high_only(self):
        """Toggle the high-score (≥70%) filter on/off."""
        val = not self._screen_high_only.get()
        self._screen_high_only.set(val)
        if val:
            self._high_only_btn.config(bg=GREEN_LT, fg=GREEN)
        else:
            self._high_only_btn.config(bg=BG2, fg=TEXT)
        self._apply_screen_filter()

    def _update_screen_count(self):
        n = sum(v.get() for v in self._screen_vars.values())
        self._screen_status_label.config(text=f"{n} selected")

    def _screen_select_all(self):
        for var in self._screen_vars.values():
            var.set(True)
        self._update_screen_count()

    def _screen_deselect_all(self):
        for var in self._screen_vars.values():
            var.set(False)
        self._update_screen_count()

    def _score_selected(self):
        """Mark ticked discovered jobs as score_me and run scoring."""
        selected_ids = [jid for jid, var in self._screen_vars.items() if var.get()]
        if not selected_ids:
            messagebox.showinfo("Nothing selected", "Tick at least one job to score.")
            return

        # Only score discovered jobs
        to_score = []
        for job_id in selected_ids:
            job = self.tracker.get_job(job_id)
            if job and job["status"] == "discovered":
                to_score.append(job_id)

        if not to_score:
            messagebox.showinfo("Nothing to score",
                "Selected jobs have already been scored or are filtered.\n"
                "Only tick jobs with blue badges to score them.")
            return

        cost_est = len(to_score) * 0.01
        if not messagebox.askyesno(
            "Score jobs",
            f"Score {len(to_score)} job(s)?\n\n"
            f"Estimated cost: ~${cost_est:.2f} in Claude API credits.\n\n"
            f"Jobs scoring below {self.config.min_match_score}% will be filtered automatically."
        ):
            return

        for job_id in to_score:
            self.tracker.update_status(job_id, "score_me", "Selected for scoring in UI")

        self._last_scored_ids = to_score  # remember which jobs were in this batch
        self._score_btn.config(text="Scoring…", state="disabled", bg="#555")
        self.update_idletasks()

        self._scan_log_lines.append(
            f"── Scoring started {datetime.now().strftime('%d %b %Y %H:%M:%S')} ──"
        )

        def score_thread():
            capture = _LogCapture(sys.__stdout__, on_line=self._append_log_line)
            old_stdout = sys.stdout
            sys.stdout = capture
            try:
                from main import run_score_selected
                run_score_selected(self.config, self.tracker)
            except Exception as e:
                err = str(e)
                self._append_log_line(f"ERROR: {err}")
                self.after(0, lambda: messagebox.showerror("Scoring failed", err))
            finally:
                sys.stdout = old_stdout
                self._append_log_line(
                    f"── Scoring finished {datetime.now().strftime('%H:%M:%S')} ──"
                )
                self.after(0, self._score_done)

        threading.Thread(target=score_thread, daemon=True).start()

    def _score_done(self):
        self._score_btn.config(text="📊  Score selected (~$0.01 each)",
                               state="normal", bg=BLUE)
        self._refresh_screen()
        batch_ids = set(getattr(self, "_last_scored_ids", []))
        batch_jobs = [self.tracker.get_job(jid) for jid in batch_ids if self.tracker.get_job(jid)]
        self._session.record_scored(batch_jobs)
        scored   = [j for j in batch_jobs if j["status"] == "scored"]
        filtered = [j for j in batch_jobs if j["status"] == "filtered"]
        if scored or filtered:
            min_score = getattr(self.config, "min_match_score", 65) if self.config else 65
            high   = sum(1 for j in scored if j.get("match_score", 0) >= 70)
            medium = sum(1 for j in scored if min_score <= j.get("match_score", 0) < 70)
            low    = len(filtered)
            lines = [f"Scoring complete — {len(scored)} job(s) passed threshold.\n"]
            if high:
                lines.append(f"  ✓  {high} high match (≥70%) — strong candidates")
            if medium:
                lines.append(f"  ~  {medium} medium match ({min_score}–69%) — worth considering")
            if low:
                lines.append(f"  ✗  {low} low score (<{min_score}%) — auto-hidden")
            lines.append("\nReview scores below. Tick jobs to tailor and click 'Tailor scored jobs'.")
            messagebox.showinfo("Scoring complete", "\n".join(lines))

    def _tailor_selected(self):
        selected_ids = [jid for jid, var in self._screen_vars.items() if var.get()]
        if not selected_ids:
            messagebox.showinfo("Nothing selected", "Tick at least one job to tailor.")
            return

        # Only tailor scored jobs
        to_tailor = []
        for job_id in selected_ids:
            job = self.tracker.get_job(job_id)
            if job and job["status"] == "scored":
                to_tailor.append(job_id)

        if not to_tailor:
            messagebox.showinfo("Nothing to tailor",
                "Please score jobs first before tailoring.\n"
                "Tick discovered jobs and click 'Score selected' first.")
            return

        cost_est = len(to_tailor) * 0.06
        if not messagebox.askyesno(
            "Tailor applications",
            f"Tailor {len(to_tailor)} application(s)?\n\n"
            f"Estimated cost: ~${cost_est:.2f} in Claude API credits.\n\n"
            f"CV and cover letter will be generated for each."
        ):
            return

        for job_id in to_tailor:
            self.tracker.update_status(job_id, "tailoring", "Selected in screening UI")
        selected_ids = to_tailor
        self._last_tailored_ids = to_tailor

        self._tailor_btn.config(text="Tailoring…", state="disabled", bg="#555")
        self.update_idletasks()

        self._scan_log_lines.append(
            f"── Tailoring started {datetime.now().strftime('%d %b %Y %H:%M:%S')} ──"
        )

        def tailor_thread():
            capture = _LogCapture(sys.__stdout__, on_line=self._append_log_line)
            old_stdout = sys.stdout
            sys.stdout = capture
            try:
                from main import run_tailor_approved
                run_tailor_approved(self.config, self.tracker)
            except Exception as e:
                err = str(e)
                self._append_log_line(f"ERROR: {err}")
                self.after(0, lambda: messagebox.showerror("Tailoring failed", err))
            finally:
                sys.stdout = old_stdout
                self._append_log_line(
                    f"── Tailoring finished {datetime.now().strftime('%H:%M:%S')} ──"
                )
                self.after(0, self._tailor_done)

        threading.Thread(target=tailor_thread, daemon=True).start()

    def _tailor_done(self):
        self._tailor_btn.config(text="✍  Tailor selected jobs", state="normal", bg=BLUE)
        self._refresh_screen()
        self._refresh_review()
        tailored_ids = set(getattr(self, "_last_tailored_ids", []))
        tailored_jobs = [self.tracker.get_job(jid) for jid in tailored_ids if self.tracker.get_job(jid)]
        self._session.record_tailored(tailored_jobs)
        n = len(self.tracker.get_pending_review()) if self.tracker else 0
        if n:
            messagebox.showinfo("Done", f"{n} application(s) tailored.\n\nGo to Review queue to check them.")
            self._show_screen("review")

    # ── REVIEW SCREEN ─────────────────────────────────────────────────────────

    def _build_review_screen(self, parent: tk.Frame):
        self._review_jobs = []
        self._review_idx = 0

        # Header
        hdr = tk.Frame(parent, bg=BG)
        hdr.pack(fill="x", padx=32, pady=(28, 0))
        tk.Label(hdr, text="Review queue", font=FONT_HEAD, bg=BG, fg=TEXT).pack(side="left")
        self._review_counter_label = tk.Label(hdr, text="", font=FONT_SM, bg=BG, fg=TEXT2)
        self._review_counter_label.pack(side="left", padx=16)

        # Job card
        self._review_card = tk.Frame(parent, bg=CARD, relief="flat", bd=1)
        self._review_card.pack(fill="x", padx=32, pady=16)

        self._rv_title     = tk.Label(self._review_card, text="", font=(FONT_FAMILY, 15, "bold"),
                                      bg=CARD, fg=TEXT, anchor="w")
        self._rv_title.pack(fill="x", padx=20, pady=(16, 2))
        self._rv_employer  = tk.Label(self._review_card, text="", font=FONT,
                                      bg=CARD, fg=TEXT2, anchor="w")
        self._rv_employer.pack(fill="x", padx=20)
        self._rv_meta      = tk.Label(self._review_card, text="", font=FONT_SM,
                                      bg=CARD, fg=TEXT2, anchor="w")
        self._rv_meta.pack(fill="x", padx=20, pady=(2, 4))
        self._rv_url       = tk.Label(self._review_card, text="", font=(FONT_FAMILY, 11, "underline"),
                                      bg=CARD, fg=BLUE, anchor="w", cursor="hand2")
        self._rv_url.pack(fill="x", padx=20, pady=(0, 4))
        self._rv_url.bind("<Button-1>", lambda e: self._open_review_url())
        self._rv_reason    = tk.Label(self._review_card, text="", font=FONT_SM,
                                      bg=CARD, fg=TEXT2, anchor="w", wraplength=700, justify="left")
        self._rv_reason.pack(fill="x", padx=20, pady=(0, 12))

        # File buttons
        btn_row = tk.Frame(self._review_card, bg=CARD)
        btn_row.pack(fill="x", padx=20, pady=(0, 16))
        tk.Button(btn_row, text="Open CV (.docx)", font=FONT_SM,
                  bg=BLUE_LT, fg=BLUE, relief="flat", padx=12, pady=6,
                  command=lambda: self._open_review_file("cv")).pack(side="left", padx=(0, 8))
        tk.Button(btn_row, text="Open cover letter (.docx)", font=FONT_SM,
                  bg=BLUE_LT, fg=BLUE, relief="flat", padx=12, pady=6,
                  command=lambda: self._open_review_file("letter")).pack(side="left")

        # Notes
        tk.Label(parent, text="Notes (optional):", font=FONT_SM, bg=BG, fg=TEXT2).pack(anchor="w", padx=32)
        self._rv_notes = tk.Text(parent, height=3, font=FONT_SM, bg=CARD,
                                 relief="solid", bd=1, fg=TEXT)
        self._rv_notes.pack(fill="x", padx=32, pady=(2, 12))

        # Decision buttons
        dec_row = tk.Frame(parent, bg=BG)
        dec_row.pack(fill="x", padx=32)

        tk.Button(dec_row, text="← Previous", font=FONT_SM,
                  bg=BG2, fg=TEXT, relief="flat", padx=14, pady=8,
                  command=self._review_prev).pack(side="left")

        tk.Button(dec_row, text="Skip this application", font=FONT,
                  bg=RED_LT, fg=RED, relief="flat", padx=16, pady=8,
                  command=self._review_skip).pack(side="left", padx=(0, 8))

        tk.Button(dec_row, text="✓ Approve", font=(FONT_FAMILY, 13, "bold"),
                  bg=GREEN, fg="white", relief="flat", padx=20, pady=8,
                  command=self._review_approve).pack(side="left")

        tk.Button(dec_row, text="Next →", font=FONT_SM,
                  bg=BG2, fg=TEXT, relief="flat", padx=14, pady=8,
                  command=self._review_next).pack(side="right")

        # Auto-apply row — button shown/hidden based on job source
        self._apply_row = tk.Frame(parent, bg=BG)
        self._apply_row.pack(fill="x", padx=32, pady=(12, 0))
        self._auto_apply_btn = tk.Button(
            self._apply_row,
            text="",
            font=(FONT_FAMILY, 13, "bold"),
            relief="flat", padx=20, pady=8,
            command=self._review_auto_apply,
        )
        self._auto_apply_btn.pack(side="left")
        self._auto_apply_status = tk.Label(
            self._apply_row, text="", font=FONT_SM, bg=BG, fg=TEXT2
        )
        self._auto_apply_status.pack(side="left", padx=12)
        self._apply_row.pack_forget()  # hidden until a supported job is shown

    def _refresh_review(self):
        if not self.tracker:
            return

        # Show splash while loading
        self._rv_title.config(text="⏳  Refreshing…")
        self._rv_employer.config(text="")
        self._rv_meta.config(text="")
        self._rv_url.config(text="")
        self._rv_reason.config(text="")
        self._review_counter_label.config(text="")
        self.update_idletasks()

        self._review_jobs = self.tracker.get_pending_review()
        self._review_idx = 0
        self._show_review_job()

    def _show_review_job(self):
        if not self._review_jobs:
            self._rv_title.config(text="No applications pending review")
            self._rv_employer.config(text="Run 'Scan for jobs' to find new positions.")
            self._rv_meta.config(text="")
            self._rv_url.config(text="")
            self._rv_reason.config(text="")
            self._review_counter_label.config(text="")
            return

        if self._review_idx >= len(self._review_jobs):
            self._review_idx = len(self._review_jobs) - 1

        job = self._review_jobs[self._review_idx]
        n = len(self._review_jobs)
        self._review_counter_label.config(text=f"{self._review_idx + 1} of {n}")

        score = f"{job['match_score']}%" if job.get("match_score") else "N/A"
        source = (job.get("source") or "").replace("_", " ").title()
        salary = job.get("salary") or "Salary not listed"
        location = job.get("location") or ""
        closes = job.get("date_closes") or ""
        closes_str = f" · Closes: {closes}" if closes else ""

        self._rv_title.config(text=job.get("title") or "Unknown role")
        self._rv_employer.config(text=job.get("employer") or "")
        self._rv_meta.config(text=f"{location}  ·  {salary}  ·  Match: {score}  ·  Source: {source}{closes_str}")

        url = job.get("url") or ""
        self._rv_url.config(text=url if url else "")
        self._rv_url._job_url = url

        self._rv_reason.config(text=job.get("match_reason") or "")
        self._rv_notes.delete("1.0", "end")
        if job.get("notes"):
            self._rv_notes.insert("1.0", job["notes"])

        # Show auto-apply button for supported sources
        src = (job.get("source") or "").lower()
        apply_label = None
        apply_colours = None
        if src == "reed":
            apply_label = "Approve & Apply on Reed"
            apply_colours = ("#8B2FC9", "#F3E8FF")   # purple
        elif src in ("linkedin", "linkedin_rss", "linkedin_manual"):
            easy_apply = self.tracker.is_easy_apply(job) if self.tracker else None
            if easy_apply is not False:  # True or None (unknown) — show button
                apply_label = "Approve & Easy Apply on LinkedIn"
                apply_colours = (BLUE, BLUE_LT)

        if apply_label:
            self._auto_apply_btn.config(
                text=apply_label,
                bg=apply_colours[0],
                fg=TEXT,
            )
            self._auto_apply_status.config(text="")
            self._apply_row.pack(fill="x", padx=32, pady=(12, 0))
        else:
            self._apply_row.pack_forget()

    def _open_review_url(self):
        url = getattr(self._rv_url, "_job_url", "")
        if url:
            webbrowser.open(url)

    def _open_review_file(self, which: str):
        if not self._review_jobs or self._review_idx >= len(self._review_jobs):
            return
        job = self._review_jobs[self._review_idx]
        path = None
        if which == "cv":
            path = job.get("tailored_cv_path")
        elif which == "letter":
            path = job.get("cover_letter_path")
        elif which == "cv_pdf":
            base = job.get("tailored_cv_path") or ""
            path = base.replace(".docx", ".pdf").replace(".md", ".pdf")
        elif which == "letter_pdf":
            base = job.get("cover_letter_path") or ""
            path = base.replace(".docx", ".pdf").replace(".md", ".pdf")
        if path and Path(path).exists():
            self._open_file(path)
        else:
            messagebox.showinfo("File not found", f"The file does not exist yet:\n{path}")

    def _save_review_notes(self):
        if not self._review_jobs:
            return
        job = self._review_jobs[self._review_idx]
        notes = self._rv_notes.get("1.0", "end").strip()
        if notes:
            self.tracker.add_note(job["id"], notes)

    def _review_approve(self):
        if not self._review_jobs:
            return
        self._save_review_notes()
        job = self._review_jobs[self._review_idx]
        self.tracker.update_status(job["id"], "approved", "Approved via UI review")
        self._review_jobs.pop(self._review_idx)
        if self._review_idx >= len(self._review_jobs):
            self._review_idx = max(0, len(self._review_jobs) - 1)
        n = len(self._review_jobs)
        self._review_counter_label.config(text=f"{self._review_idx + 1 if n else 0} of {n}")
        self._show_review_job()

    def _review_auto_apply(self):
        """Approve the current job and launch the appropriate auto-apply module in a thread."""
        if not self._review_jobs:
            return
        self._save_review_notes()
        job = self._review_jobs[self._review_idx]
        src = (job.get("source") or "").lower()

        # Warn if Easy Apply status is unknown for LinkedIn jobs
        if src in ("linkedin", "linkedin_rss", "linkedin_manual"):
            easy_apply = self.tracker.is_easy_apply(job) if self.tracker else None
            if easy_apply is None:
                proceed = messagebox.askyesno(
                    "Easy Apply not confirmed",
                    "This job was scanned before Easy Apply detection was added, "
                    "so it's not known whether it supports LinkedIn Easy Apply.\n\n"
                    "If the job doesn't have Easy Apply, the browser will open "
                    "but won't be able to submit automatically.\n\n"
                    "Try anyway?",
                )
                if not proceed:
                    return

        self._auto_apply_btn.config(state="disabled")
        self._auto_apply_status.config(text="Opening browser…", fg=TEXT2)

        def run():
            try:
                if src == "reed":
                    from reed_apply import ReedApplicant
                    applicant = ReedApplicant(self.config, headless=False)
                    submitted = applicant.apply(job)
                elif src in ("linkedin", "linkedin_rss", "linkedin_manual"):
                    from linkedin_apply import LinkedInApplicant
                    applicant = LinkedInApplicant(self.config, headless=False)
                    submitted = applicant.apply(job)
                else:
                    submitted = False

                if submitted:
                    self.tracker.update_status(job["id"], "approved", "Approved via UI review")
                    self.tracker.update_status(job["id"], "submitted", f"Auto-submitted via {src} apply")
                    self.after(0, lambda: self._auto_apply_done(job, platform=src, success=True))
                else:
                    self.after(0, lambda: self._auto_apply_done(job, platform=src, success=False))
            except Exception as e:
                err = str(e)
                self.after(0, lambda: self._auto_apply_done(job, platform=src, success=False, error=err))

        threading.Thread(target=run, daemon=True).start()

    def _auto_apply_done(self, job: dict, success: bool, platform: str = "", error: str = ""):
        self._session.record_apply(job, platform or "unknown", success)
        self._auto_apply_btn.config(state="normal")
        if success:
            self._auto_apply_status.config(text="✓ Submitted!", fg=GREEN)
            self._review_jobs.pop(self._review_idx)
            if self._review_idx >= len(self._review_jobs):
                self._review_idx = max(0, len(self._review_jobs) - 1)
            self._show_review_job()
        elif error:
            self._auto_apply_status.config(text=f"Error: {error}", fg=RED)
        else:
            self._auto_apply_status.config(
                text="No Easy Apply found — apply manually on LinkedIn.", fg=AMBER
            )

    def _review_skip(self):
        if not self._review_jobs:
            return
        self._save_review_notes()
        job = self._review_jobs[self._review_idx]
        self.tracker.update_status(job["id"], "skipped", "Skipped via UI review")
        self._review_jobs.pop(self._review_idx)
        if self._review_idx >= len(self._review_jobs):
            self._review_idx = max(0, len(self._review_jobs) - 1)
        self._show_review_job()

    def _review_next(self):
        self._save_review_notes()
        if self._review_jobs and self._review_idx < len(self._review_jobs) - 1:
            self._review_idx += 1
            self._show_review_job()

    def _review_prev(self):
        self._save_review_notes()
        if self._review_idx > 0:
            self._review_idx -= 1
            self._show_review_job()

    # ── SETTINGS SCREEN ───────────────────────────────────────────────────────

    def _refresh_settings(self):
        """Rebuild settings screen content — called when switching to settings tab."""
        if not hasattr(self, "_settings_parent"):
            return
        # Destroy and rebuild inner content so config values are fresh
        for w in self._settings_parent.winfo_children():
            w.destroy()
        self._build_settings_screen(self._settings_parent)

    def _build_settings_screen(self, parent: tk.Frame):
        self._settings_parent = parent
        canvas = tk.Canvas(parent, bg=BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        inner = tk.Frame(canvas, bg=BG)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        tk.Label(inner, text="Settings", font=FONT_HEAD, bg=BG, fg=TEXT).pack(anchor="w", padx=40, pady=(32, 4))
        tk.Label(inner, text="Changes take effect on the next scan. Restart the app after saving.",
                 font=FONT_SM, bg=BG, fg=TEXT2).pack(anchor="w", padx=40, pady=(0, 8))

        # ── Job search ────────────────────────────────────────────────────────
        self._section_label(inner, "Job search")

        tk.Label(inner, text="Search keywords (one per line):", font=FONT_SM, bg=BG, fg=TEXT2).pack(anchor="w", padx=40, pady=(8, 2))
        self._kw_text = tk.Text(inner, height=7, font=FONT_SM, bg=CARD,
                                relief="solid", bd=1, fg=TEXT, width=52)
        self._kw_text.pack(anchor="w", padx=40)
        if self.config:
            self._kw_text.insert("1.0", "\n".join(self.config.search_keywords))

        self._min_salary_var = self._labeled_entry(inner, "Minimum salary (£)", "20000")
        self._max_salary_var = self._labeled_entry(inner, "Maximum salary — 0 = no limit (£)", "0")
        self._max_age_var    = self._labeled_entry(inner, "Only show jobs posted within (days)", "14")
        self._min_score_var  = self._labeled_entry(inner, "Minimum match score to keep (%)", "75")

        if self.config:
            self._min_salary_var.set(str(self.config.min_salary_gbp))
            self._max_salary_var.set(str(getattr(self.config, 'max_salary_gbp', 0)))
            self._max_age_var.set(str(self.config.max_job_age_days))
            self._min_score_var.set(str(self.config.min_match_score))

        # ATS mode toggle
        self._ats_mode_var = tk.BooleanVar(value=self.config.ats_mode if self.config else True)
        ats_row = tk.Frame(inner, bg=BG)
        ats_row.pack(anchor="w", padx=40, pady=(8, 0))
        tk.Checkbutton(ats_row, text="ATS compatibility mode",
                       variable=self._ats_mode_var, font=FONT, bg=BG, fg=TEXT,
                       activebackground=BG, selectcolor=CARD).pack(side="left")
        tk.Label(ats_row, text="  (recommended — makes CVs readable by automated scanners)",
                 font=FONT_SM, bg=BG, fg=TEXT2).pack(side="left")

        # ── Job boards ────────────────────────────────────────────────────────
        self._section_label(inner, "Job boards")

        self._scan_civil_var      = tk.BooleanVar(value=self.config.scan_civil_service if self.config else True)
        self._scan_guardian_var   = tk.BooleanVar(value=self.config.scan_guardian if self.config else True)
        self._scan_linkedin_var   = tk.BooleanVar(value=self.config.scan_linkedin if self.config else True)
        self._scan_w4mpjobs_var   = tk.BooleanVar(value=self.config.scan_w4mpjobs if self.config else True)
        self._scan_charityjob_var = tk.BooleanVar(value=self.config.scan_charityjob if self.config else True)
        for var, label in [
            (self._scan_civil_var,      "Civil Service Jobs"),
            (self._scan_guardian_var,   "Guardian Jobs"),
            (self._scan_linkedin_var,   "LinkedIn (RSS feed)"),
            (self._scan_w4mpjobs_var,   "W4MP Jobs (parliamentary & political roles)"),
            (self._scan_charityjob_var, "Charity Job"),
        ]:
            tk.Checkbutton(inner, text=label, variable=var, font=FONT, bg=BG, fg=TEXT,
                           activebackground=BG, selectcolor=CARD).pack(anchor="w", padx=40, pady=2)

        # ── Filter editor ─────────────────────────────────────────────────────
        self._section_label(inner, "Filters — Too senior (title keywords)")
        tk.Label(inner,
                 text="Jobs whose title contains any of these words are removed before scoring (free). Add words to reduce volume. Remove words if good roles are being filtered out.",
                 font=FONT_SM, bg=BG, fg=TEXT2, justify="left").pack(anchor="w", padx=40, pady=(4, 4))
        self._senior_filter_frame = self._build_filter_editor(
            inner, "seniority_filter_titles",
            self.config.seniority_filter_titles if self.config else []
        )

        self._section_label(inner, "Filters — Irrelevant job titles")
        tk.Label(inner,
                 text="Jobs whose title contains any of these words are removed as clearly irrelevant. Add words to remove more noise. Remove words if relevant roles are being missed.",
                 font=FONT_SM, bg=BG, fg=TEXT2, justify="left").pack(anchor="w", padx=40, pady=(4, 4))
        self._irrelevant_filter_frame = self._build_filter_editor(
            inner, "irrelevant_filter_titles",
            self.config.irrelevant_filter_titles if self.config else []
        )

        self._section_label(inner, "Filters — Excluded locations")
        tk.Label(inner,
                 text="Jobs whose location contains any of these words are removed. Add cities/countries to exclude. Remove entries to include more locations.",
                 font=FONT_SM, bg=BG, fg=TEXT2, justify="left").pack(anchor="w", padx=40, pady=(4, 4))
        self._location_filter_frame = self._build_filter_editor(
            inner, "exclude_locations",
            self.config.exclude_locations if self.config else []
        )

        # ── Save ──────────────────────────────────────────────────────────────
        tk.Frame(inner, bg=BG, height=16).pack()
        save_row = tk.Frame(inner, bg=BG)
        save_row.pack(anchor="w", padx=40, pady=(0, 40))
        tk.Button(save_row, text="Save all settings", font=(FONT_FAMILY, 13, "bold"),
                  bg=BLUE, fg=TEXT,  relief="flat", padx=20, pady=8,
                  command=self._save_settings).pack(side="left", padx=(0, 12))
        tk.Button(save_row, text="Reset filters to defaults", font=FONT_SM,
                  bg=BG2, fg=TEXT, relief="flat", padx=12, pady=8,
                  command=self._reset_filters).pack(side="left")

    def _build_filter_editor(self, parent, config_attr: str, initial_values: list) -> tk.Frame:
        """Build an editable list widget for a filter keyword list."""
        outer = tk.Frame(parent, bg=BG)
        outer.pack(anchor="w", padx=40, pady=(0, 8), fill="x")

        # Listbox showing current keywords
        list_frame = tk.Frame(outer, bg=BG)
        list_frame.pack(side="left", fill="both")

        listbox = tk.Listbox(list_frame, font=FONT_SM, height=6, width=38,
                             bg=CARD, fg=TEXT, relief="solid", bd=1,
                             selectbackground=BLUE_LT, selectforeground=BLUE,
                             activestyle="none")
        listbox.pack(side="left", fill="both")
        sb = ttk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
        sb.pack(side="right", fill="y")
        listbox.config(yscrollcommand=sb.set)

        for val in initial_values:
            listbox.insert("end", val)

        # Controls to the right
        ctrl = tk.Frame(outer, bg=BG)
        ctrl.pack(side="left", padx=(12, 0), anchor="n")

        entry = tk.Entry(ctrl, font=FONT_SM, width=22, relief="solid", bd=1, bg=CARD)
        entry.pack(pady=(0, 6))
        entry.bind("<Return>", lambda e: self._filter_add(listbox, entry))

        tk.Button(ctrl, text="+ Add", font=FONT_SM, bg=GREEN_LT, fg=GREEN,
                  relief="flat", padx=10, pady=4, width=10,
                  command=lambda: self._filter_add(listbox, entry)).pack(pady=(0, 4))

        tk.Button(ctrl, text="Remove selected", font=FONT_SM, bg=RED_LT, fg=RED,
                  relief="flat", padx=10, pady=4, width=14,
                  command=lambda: self._filter_remove(listbox)).pack(pady=(0, 4))

        tk.Label(ctrl, text="Double-click to edit", font=(FONT_FAMILY, 10),
                 bg=BG, fg=TEXT2).pack()

        listbox.bind("<Double-1>", lambda e: self._filter_edit(listbox, entry))

        # Store reference for save
        listbox._config_attr = config_attr
        outer._listbox = listbox
        outer._config_attr = config_attr
        return outer

    @staticmethod
    def _filter_add(listbox: tk.Listbox, entry: tk.Entry):
        val = entry.get().strip().lower()
        if val and val not in listbox.get(0, "end"):
            listbox.insert("end", val)
            entry.delete(0, "end")

    @staticmethod
    def _filter_remove(listbox: tk.Listbox):
        sel = listbox.curselection()
        for idx in reversed(sel):
            listbox.delete(idx)

    @staticmethod
    def _filter_edit(listbox: tk.Listbox, entry: tk.Entry):
        sel = listbox.curselection()
        if not sel:
            return
        val = listbox.get(sel[0])
        entry.delete(0, "end")
        entry.insert(0, val)
        listbox.delete(sel[0])

    def _save_settings(self):
        """Write all settings back to config.py."""
        if not self.config:
            return

        try:
            config_path = Path(__file__).parent / "config.py"
            text = config_path.read_text(encoding="utf-8")

            def set_int(attr, val):
                import re
                return re.sub(
                    rf"({attr}:\s*int\s*=\s*)\d+",
                    rf"\g<1>{val}", text
                )
            def set_bool(attr, val):
                import re
                return re.sub(
                    rf"({attr}:\s*bool\s*=\s*)(?:True|False)",
                    rf"\g<1>{'True' if val else 'False'}", text
                )
            def set_list(attr, values):
                import re
                items = ",\n        ".join(f'"{v}"' for v in values)
                new_block = f'{attr}: List[str] = field(default_factory=lambda: [\n        {items},\n    ])'
                return re.sub(
                    rf'{attr}: List\[str\] = field\(default_factory=lambda: \[.*?\]\)',
                    new_block, text, flags=re.DOTALL
                )

            # Scalar settings
            text = set_int("min_salary_gbp", self._min_salary_var.get() or "20000")
            text = set_int("max_salary_gbp", self._max_salary_var.get() or "0")
            text = set_int("max_job_age_days", self._max_age_var.get() or "14")
            text = set_int("min_match_score", self._min_score_var.get() or "75")
            text = set_bool("ats_mode", self._ats_mode_var.get())
            text = set_bool("scan_civil_service", self._scan_civil_var.get())
            text = set_bool("scan_guardian", self._scan_guardian_var.get())
            text = set_bool("scan_linkedin", self._scan_linkedin_var.get())
            text = set_bool("scan_w4mpjobs", self._scan_w4mpjobs_var.get())
            text = set_bool("scan_charityjob", self._scan_charityjob_var.get())

            # Keywords
            keywords = [k.strip() for k in self._kw_text.get("1.0", "end").splitlines() if k.strip()]
            if keywords:
                text = set_list("search_keywords", keywords)

            # Filter lists
            for frame in [self._senior_filter_frame, self._irrelevant_filter_frame, self._location_filter_frame]:
                lb = frame._listbox
                values = list(lb.get(0, "end"))
                text = set_list(lb._config_attr, values)

            config_path.write_text(text, encoding="utf-8")

            # Reload config in memory
            self._load_backend()

            messagebox.showinfo("Settings saved",
                "✓ Settings saved successfully.\n\n"
                "Filter changes take effect immediately on the next scan.\n"
                "Restart the app to reload all other settings.")

        except Exception as e:
            messagebox.showerror("Save failed", f"Could not save settings:\n{e}")

    def _reset_filters(self):
        """Reset all three filter lists to their default values."""
        if not messagebox.askyesno("Reset filters",
            "Reset seniority, irrelevant title, and location filters to defaults?\n\n"
            "Your current filter changes will be lost."):
            return

        from config import Config
        defaults = Config()

        for frame, attr in [
            (self._senior_filter_frame,     "seniority_filter_titles"),
            (self._irrelevant_filter_frame,  "irrelevant_filter_titles"),
            (self._location_filter_frame,    "exclude_locations"),
        ]:
            lb = frame._listbox
            lb.delete(0, "end")
            for val in getattr(defaults, attr, []):
                lb.insert("end", val)

        messagebox.showinfo("Filters reset", "Filters reset to defaults. Click Save all settings to apply.")

    # ── SCAN ──────────────────────────────────────────────────────────────────

    # ── SCAN LOG SCREEN ───────────────────────────────────────────────────────

    def _build_log_screen(self, parent: tk.Frame):
        header = tk.Frame(parent, bg=BG)
        header.pack(fill="x", padx=32, pady=(28, 8))
        tk.Label(header, text="Scan Log", font=FONT_HEAD, bg=BG, fg=TEXT).pack(side="left")

        btn_row = tk.Frame(header, bg=BG)
        btn_row.pack(side="right")
        tk.Button(btn_row, text="Clear", font=FONT_SM, bg=BG2, fg=TEXT,
                  relief="flat", padx=10, pady=4,
                  command=self._clear_log).pack(side="left", padx=(0, 8))
        tk.Button(btn_row, text="Copy all", font=FONT_SM, bg=BG2, fg=TEXT,
                  relief="flat", padx=10, pady=4,
                  command=self._copy_log).pack(side="left")

        tk.Label(parent,
                 text="Shows output from scans, scoring and tailoring — errors, jobs found, jobs filtered.",
                 font=FONT_SM, bg=BG, fg=TEXT2).pack(anchor="w", padx=32, pady=(0, 8))

        self._log_text = scrolledtext.ScrolledText(
            parent, font=("Menlo" if sys.platform == "darwin" else "Consolas", 11),
            bg="#1E1E1E", fg="#D4D4D4", insertbackground="#D4D4D4",
            relief="flat", bd=0, wrap="word", state="disabled",
        )
        self._log_text.pack(fill="both", expand=True, padx=32, pady=(0, 24))

        # Colour tags
        self._log_text.tag_configure("error",    foreground="#F48771")
        self._log_text.tag_configure("ok",        foreground="#89D185")
        self._log_text.tag_configure("filtered",  foreground="#808080")
        self._log_text.tag_configure("heading",   foreground="#569CD6")
        self._log_text.tag_configure("warn",      foreground="#CCA700")

        self._refresh_log_screen()

    def _log_line_tag(self, line: str) -> str:
        low = line.lower()
        if any(k in low for k in ("error", "failed", "exception", "traceback", "credit", "billing")):
            return "error"
        if any(k in low for k in ("✓", "saved", "scored", "tailored", "found")):
            return "ok"
        if any(k in low for k in ("⊘", "filtered", "senior", "irrelevant", "location", "salary", "skipped")):
            return "filtered"
        if any(k in low for k in ("scanning", "---", "===", "phase", "complete")):
            return "heading"
        if any(k in low for k in ("warning", "warn", "no jobs", "0 jobs", "blocked")):
            return "warn"
        return ""

    def _refresh_log_screen(self):
        if not hasattr(self, "_log_text"):
            return
        self._log_text.config(state="normal")
        self._log_text.delete("1.0", "end")
        if not self._scan_log_lines:
            self._log_text.insert("end", "No scan has been run yet.\nClick 'Scan for jobs' to start.\n",
                                  "filtered")
        else:
            for line in self._scan_log_lines:
                tag = self._log_line_tag(line)
                self._log_text.insert("end", line + "\n", tag)
            self._log_text.see("end")
        self._log_text.config(state="disabled")

    def _append_log_line(self, line: str):
        """Called from scan thread — schedules a UI update on the main thread."""
        self._scan_log_lines.append(line)
        if hasattr(self, "_log_text"):
            def _update():
                self._log_text.config(state="normal")
                tag = self._log_line_tag(line)
                self._log_text.insert("end", line + "\n", tag)
                self._log_text.see("end")
                self._log_text.config(state="disabled")
            self.after(0, _update)

    def _clear_log(self):
        self._scan_log_lines.clear()
        self._refresh_log_screen()

    def _copy_log(self):
        text = "\n".join(self._scan_log_lines)
        self.clipboard_clear()
        self.clipboard_append(text)
        messagebox.showinfo("Copied", "Log copied to clipboard.")

    # ── INFO SCREEN ──────────────────────────────────────────────────────────────

    def _build_info_screen(self, parent: tk.Frame):
        tk.Label(parent, text="Info & Documentation", font=FONT_HEAD, bg=BG, fg=TEXT).pack(
            anchor="w", padx=40, pady=(32, 4))
        tk.Label(parent, text="Open project documents from the docs/ folder.",
                 font=FONT_SM, bg=BG, fg=TEXT2).pack(anchor="w", padx=40, pady=(0, 24))

        docs = [
            ("Business Requirements Document (BRD)", "job_agent_BRD.docx",
             "Full specification of what the Job Agent does and why."),
            ("User Guide", "job_agent_user_guide.docx",
             "Step-by-step instructions for using the application."),
        ]
        for title, filename, desc in docs:
            card = tk.Frame(parent, bg=CARD, relief="flat", bd=1)
            card.pack(fill="x", padx=40, pady=(0, 12))
            tk.Label(card, text=title, font=(FONT_FAMILY, 13, "bold"),
                     bg=CARD, fg=TEXT).pack(anchor="w", padx=20, pady=(14, 2))
            tk.Label(card, text=desc, font=FONT_SM, bg=CARD, fg=TEXT2).pack(
                anchor="w", padx=20, pady=(0, 10))
            tk.Button(card, text=f"Open {filename}", font=FONT_SM,
                      bg=BLUE_LT, fg=BLUE, relief="flat", padx=12, pady=6,
                      command=lambda f=filename: self._open_doc(f)).pack(
                anchor="w", padx=20, pady=(0, 14))

    def _open_user_guide(self):
        """Open the User Guide docx."""
        self._open_doc("job_agent_user_guide.docx")

    def _open_brd(self):
        """Open the BRD docx."""
        self._open_doc("job_agent_BRD.docx")

    def _open_doc(self, filename: str):
        """Open a documentation file from the docs/ folder."""
        import sys as _sys
        if getattr(_sys, "frozen", False):
            # PyInstaller bundle: docs are alongside the executable
            docs_dir = Path(_sys._MEIPASS) / "docs"
        else:
            docs_dir = Path(__file__).parent.parent / "docs"
        path = docs_dir / filename
        if not path.exists():
            messagebox.showwarning("File not found",
                f"{filename} not found.\n\nExpected location:\n{path}")
            return
        self._open_file(str(path))

    def _run_refilter(self):
        """Re-apply filters to discovered jobs without rescanning. FREE."""
        if not self.config or not self.tracker:
            return

        def refilter_thread():
            try:
                from main import run_refilter
                run_refilter(self.config, self.tracker)
            except Exception as e:
                self.after(0, lambda: messagebox.showerror("Refilter failed", str(e)))
            finally:
                self.after(0, self._refresh_screen)
                self.after(0, self._refresh_dashboard)

        threading.Thread(target=refilter_thread, daemon=True).start()
        self._show_screen("screen")
        messagebox.showinfo("Refiltering", "Re-applying filters to discovered jobs...\nThe list will update when done.")

    def _run_scan(self):
        if self._needs_setup():
            messagebox.showwarning("Setup required", "Please complete setup before scanning.")
            self._show_screen("setup")
            return

        self._scan_btn.config(text="  Scanning…", state="disabled", bg="#555")
        self.update_idletasks()

        self._scan_log_lines.append(
            f"── Scan started {datetime.now().strftime('%d %b %Y %H:%M:%S')} ──"
        )

        def scan_thread():
            capture = _LogCapture(sys.__stdout__, on_line=self._append_log_line)
            old_stdout = sys.stdout
            sys.stdout = capture
            try:
                from job_scanner import JobScanner
                from cv_tailor import CVTailor
                from main import run_scan
                run_scan(self.config, self.tracker)
            except Exception as e:
                err = str(e)
                self._append_log_line(f"ERROR: {err}")
                self.after(0, lambda: messagebox.showerror("Scan failed", err))
            finally:
                sys.stdout = old_stdout
                self._append_log_line(
                    f"── Scan finished {datetime.now().strftime('%H:%M:%S')} ──"
                )
                self.after(0, self._scan_done)

        threading.Thread(target=scan_thread, daemon=True).start()

    def _scan_done(self):
        self._scan_btn.config(text="  Scan for jobs", state="normal", bg=BLUE)
        self._refresh_dashboard()
        self._refresh_review()
        pending = len(self.tracker.get_pending_review()) if self.tracker else 0
        discovered = len(self.tracker.get_discovered_jobs()) if self.tracker else 0
        if discovered:
            messagebox.showinfo("Scan complete",
                                f"Scan finished.\n\n{discovered} job(s) found and ready to review.\n\n"
                                f"No credits spent yet — review titles in Screen jobs first.")
            self._show_screen("screen")


    # ── DATABASE SCREEN ───────────────────────────────────────────────────────

    def _build_database_screen(self, parent: tk.Frame):
        tk.Label(parent, text="Database", font=FONT_HEAD, bg=BG, fg=TEXT).pack(
            anchor="w", padx=32, pady=(28, 2))
        tk.Label(parent, text="Browse all records, or write SQL to query the database.",
                 font=FONT_SM, bg=BG, fg=TEXT2).pack(anchor="w", padx=32, pady=(0, 16))

        # ── Top pane: all records ─────────────────────────────────────────────
        top_frame = tk.Frame(parent, bg=BG)
        top_frame.pack(fill="both", expand=True, padx=32, pady=(0, 8))

        top_bar = tk.Frame(top_frame, bg=BG)
        top_bar.pack(fill="x", pady=(0, 6))
        tk.Label(top_bar, text="All records", font=(FONT_FAMILY, 12, "bold"),
                 bg=BG, fg=TEXT).pack(side="left")
        tk.Button(top_bar, text="Refresh", font=FONT_SM, bg=BG2, fg=TEXT,
                  relief="flat", padx=10, pady=4,
                  command=self._db_refresh_all).pack(side="right")

        all_cols = ("id", "title", "employer", "location", "salary",
                    "status", "match_score", "date_found", "date_updated", "url")
        tree_frame = tk.Frame(top_frame, bg=BG)
        tree_frame.pack(fill="both", expand=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        self._db_all_tree = ttk.Treeview(
            tree_frame, columns=all_cols, show="headings",
            yscrollcommand=vsb.set, xscrollcommand=hsb.set, height=8)
        vsb.config(command=self._db_all_tree.yview)
        hsb.config(command=self._db_all_tree.xview)

        col_widths = {"id": 40, "title": 220, "employer": 150, "location": 100,
                      "salary": 90, "status": 90, "match_score": 70,
                      "date_found": 120, "date_updated": 120, "url": 180}
        for c in all_cols:
            self._db_all_tree.heading(c, text=c.replace("_", " ").title())
            self._db_all_tree.column(c, width=col_widths.get(c, 100), anchor="w")

        self._db_all_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        # ── Middle pane: SQL editor ───────────────────────────────────────────
        mid_frame = tk.Frame(parent, bg=BG)
        mid_frame.pack(fill="x", padx=32, pady=(0, 8))

        mid_bar = tk.Frame(mid_frame, bg=BG)
        mid_bar.pack(fill="x", pady=(0, 4))
        tk.Label(mid_bar, text="SQL query", font=(FONT_FAMILY, 12, "bold"),
                 bg=BG, fg=TEXT).pack(side="left")
        tk.Button(mid_bar, text="Run  ▶", font=FONT_SM, bg=BLUE, fg="white",
                  relief="flat", padx=14, pady=4,
                  command=self._db_run_query).pack(side="right")

        self._db_sql_editor = scrolledtext.ScrolledText(
            mid_frame, height=5, font=("Courier New", 12),
            bg=CARD, fg=TEXT, insertbackground=TEXT,
            relief="flat", bd=1, wrap="none")
        self._db_sql_editor.pack(fill="x")
        self._db_sql_editor.insert("1.0", "SELECT * FROM jobs ORDER BY date_found DESC LIMIT 100;")

        # ── Bottom pane: query results ────────────────────────────────────────
        bot_frame = tk.Frame(parent, bg=BG)
        bot_frame.pack(fill="both", expand=True, padx=32, pady=(0, 24))

        bot_bar = tk.Frame(bot_frame, bg=BG)
        bot_bar.pack(fill="x", pady=(0, 6))
        self._db_result_label = tk.Label(bot_bar, text="Results", font=(FONT_FAMILY, 12, "bold"),
                                         bg=BG, fg=TEXT)
        self._db_result_label.pack(side="left")
        tk.Button(bot_bar, text="Export to Excel", font=FONT_SM, bg=GREEN_LT, fg=GREEN,
                  relief="flat", padx=12, pady=4,
                  command=self._db_export_results).pack(side="right")

        res_frame = tk.Frame(bot_frame, bg=BG)
        res_frame.pack(fill="both", expand=True)

        vsb2 = ttk.Scrollbar(res_frame, orient="vertical")
        hsb2 = ttk.Scrollbar(res_frame, orient="horizontal")
        self._db_result_tree = ttk.Treeview(
            res_frame, show="headings",
            yscrollcommand=vsb2.set, xscrollcommand=hsb2.set, height=8)
        vsb2.config(command=self._db_result_tree.yview)
        hsb2.config(command=self._db_result_tree.xview)

        self._db_result_tree.grid(row=0, column=0, sticky="nsew")
        vsb2.grid(row=0, column=1, sticky="ns")
        hsb2.grid(row=1, column=0, sticky="ew")
        res_frame.rowconfigure(0, weight=1)
        res_frame.columnconfigure(0, weight=1)

        self._db_result_cols: list = []  # current result column names
        self._db_result_rows: list = []  # current result data rows

        # populate all-records table immediately
        self._db_refresh_all()

    def _db_refresh_all(self):
        """Reload the top all-records Treeview from the database."""
        if not self.tracker:
            return
        jobs = self.tracker.get_all_jobs()
        self._db_all_tree.delete(*self._db_all_tree.get_children())
        cols = ("id", "title", "employer", "location", "salary",
                "status", "match_score", "date_found", "date_updated", "url")
        for job in jobs:
            self._db_all_tree.insert("", "end", values=[job.get(c, "") for c in cols])

    def _db_run_query(self):
        """Execute the SQL in the editor and display results in the bottom pane."""
        if not self.tracker:
            messagebox.showerror("Database", "No database connected.")
            return
        sql = self._db_sql_editor.get("1.0", "end").strip()
        if not sql:
            return
        import sqlite3
        try:
            conn = sqlite3.connect(self.tracker.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.execute(sql)
            rows = cur.fetchall()
            col_names = [d[0] for d in cur.description] if cur.description else []
            conn.close()
        except Exception as e:
            messagebox.showerror("SQL error", str(e))
            return

        # Update result Treeview
        self._db_result_cols = col_names
        self._db_result_rows = [list(r) for r in rows]

        self._db_result_tree["columns"] = col_names
        self._db_result_tree.delete(*self._db_result_tree.get_children())
        for c in col_names:
            self._db_result_tree.heading(c, text=c.replace("_", " ").title())
            self._db_result_tree.column(c, width=120, anchor="w")
        for row in self._db_result_rows:
            self._db_result_tree.insert("", "end", values=row)

        n = len(rows)
        self._db_result_label.config(
            text=f"Results — {n} row{'s' if n != 1 else ''}")

    def _db_export_results(self):
        """Export the current query results to Excel."""
        if not self._db_result_cols:
            messagebox.showinfo("Export", "Run a query first to get results.")
            return
        self._export_to_excel(
            self._db_result_cols,
            self._db_result_rows,
            default_name="db_query_export.xlsx",
        )


def main():
    app = JobAgentApp()
    app.mainloop()


if __name__ == "__main__":
    main()
