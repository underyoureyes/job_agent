"""
tests/unit/test_ui_utils.py
============================
Tests for ui.utils — open_file() and export_to_excel().
"""

import sys
import os
import pytest
from pathlib import Path
from unittest.mock import patch, MagicMock, call

# Ensure src is on the path
sys.path.insert(0, str(Path(__file__).parent.parent.parent / "src"))

from ui.utils import open_file, export_to_excel


# ── open_file ────────────────────────────────────────────────────────────────

class TestOpenFile:

    def test_open_file_mac(self):
        with patch("sys.platform", "darwin"), \
             patch("subprocess.run") as mock_run:
            open_file("/tmp/test.docx")
            mock_run.assert_called_once_with(["open", "/tmp/test.docx"])

    def test_open_file_windows(self):
        # os.startfile only exists on Windows; mock it unconditionally
        with patch("sys.platform", "win32"), \
             patch("os.startfile", create=True) as mock_startfile:
            open_file("C:\\Users\\test.docx")
            mock_startfile.assert_called_once_with("C:\\Users\\test.docx")

    def test_open_file_linux(self):
        with patch("sys.platform", "linux"), \
             patch("subprocess.run") as mock_run:
            open_file("/home/user/test.pdf")
            mock_run.assert_called_once_with(["xdg-open", "/home/user/test.pdf"])

    def test_open_file_converts_path_to_str(self):
        with patch("sys.platform", "darwin"), \
             patch("subprocess.run") as mock_run:
            open_file(Path("/tmp/test.docx"))
            mock_run.assert_called_once_with(["open", "/tmp/test.docx"])


# ── export_to_excel ───────────────────────────────────────────────────────────

class TestExportToExcel:

    def test_export_creates_valid_xlsx(self, tmp_path):
        """export_to_excel writes a valid xlsx file with correct headers."""
        pytest.importorskip("openpyxl")
        import openpyxl

        out_path = tmp_path / "test_export.xlsx"

        headers = ["ID", "Role", "Status"]
        rows = [
            [1, "Policy Analyst", "discovered"],
            [2, "Research Officer", "scored"],
        ]

        with patch("tkinter.filedialog.asksaveasfilename", return_value=str(out_path)), \
             patch("tkinter.messagebox.showinfo"), \
             patch("subprocess.run"):
            export_to_excel(headers, rows, default_name="test_export.xlsx")

        assert out_path.exists()
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active

        # Header row
        assert ws.cell(1, 1).value == "ID"
        assert ws.cell(1, 2).value == "Role"
        assert ws.cell(1, 3).value == "Status"

        # Data rows
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == "Policy Analyst"
        assert ws.cell(3, 2).value == "Research Officer"

    def test_export_no_path_chosen_does_not_raise(self):
        """If user cancels the save dialog, export_to_excel returns silently."""
        pytest.importorskip("openpyxl")

        with patch("tkinter.filedialog.asksaveasfilename", return_value=""), \
             patch("tkinter.messagebox.showinfo"):
            # Should not raise
            export_to_excel(["A", "B"], [[1, 2]], default_name="out.xlsx")

    def test_export_missing_openpyxl_shows_error(self):
        """If openpyxl is not installed, an error messagebox is shown."""
        import builtins
        real_import = builtins.__import__

        def mock_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("No module named 'openpyxl'")
            return real_import(name, *args, **kwargs)

        with patch("builtins.__import__", side_effect=mock_import), \
             patch("tkinter.messagebox.showerror") as mock_err:
            export_to_excel(["A"], [[1]], default_name="out.xlsx")
            mock_err.assert_called_once()
            assert "openpyxl" in mock_err.call_args[0][1]

    def test_export_row_count_in_message(self, tmp_path):
        """The success messagebox includes the row count."""
        pytest.importorskip("openpyxl")

        out_path = tmp_path / "count_test.xlsx"
        rows = [[i, f"Job {i}"] for i in range(5)]

        with patch("tkinter.filedialog.asksaveasfilename", return_value=str(out_path)), \
             patch("tkinter.messagebox.showinfo") as mock_info, \
             patch("subprocess.run"):
            export_to_excel(["ID", "Title"], rows, default_name="count_test.xlsx")

        info_msg = mock_info.call_args[0][1]
        assert "5" in info_msg
